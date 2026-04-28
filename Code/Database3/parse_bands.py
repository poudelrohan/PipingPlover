"""
parse_bands.py
--------------
Run the band-cell structuring step for any year 2019-2024.

Goal: for each PIPLbands cell, determine HOW MANY individual birds'
information is present and number them:
    "1) <raw text bird 1>
     2) <raw text bird 2>"

Pink = count genuinely unknowable (unreadable, too distant, vague mass note).
Otherwise, even prose/non-standard entries get "1) <raw text>" so the count
is always explicit.

Usage:
    python3 parse_bands.py          # runs all years 2019-2024
    python3 parse_bands.py 2021     # runs one year
"""

import re
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

ROOT     = Path(__file__).resolve().parents[2]
CLEAN_IN = ROOT / "Databases/Database3Clean"
REVIEW   = ROOT / "Databases/Database3BandReview"
REVIEW.mkdir(exist_ok=True)

PINK = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

# ── Per-year config ───────────────────────────────────────────────────────────
# band_col_re: regex that matches the band column header for that year
YEAR_CONFIG = {
    2019: dict(
        sheet="Form Responses 1",
        band_col_re=re.compile(r"^\d+\s+Band/Flag Codes", re.IGNORECASE),
    ),
    2020: dict(
        sheet="Focal Observations",
        band_col_re=re.compile(r"PIPL Band/Flag Codes \(point \d+\)", re.IGNORECASE),
    ),
    2021: dict(
        sheet="Focal Observations",
        band_col_re=re.compile(r"PIPL Band/Flag Codes \(point \d+\)", re.IGNORECASE),
    ),
    2022: dict(
        sheet="Focal Observations",
        band_col_re=re.compile(r"PIPL Band/Flag Codes \(point \d+\)", re.IGNORECASE),
    ),
    2023: dict(
        sheet="Focal Observations",
        band_col_re=re.compile(r"PIPL Band/Flag Codes \(point \d+\)", re.IGNORECASE),
    ),
    2024: dict(
        sheet="Focal Observations",
        band_col_re=re.compile(r"PIPL Band/Flag Codes \(point \d+\)", re.IGNORECASE),
    ),
}

# ── No-band detection ─────────────────────────────────────────────────────────
_NONE_BARE = {
    "none", "n/a", "na", "no", "nb", "nope", "0", "1", "2", "3", "-",
    "none.", "n.a.", "no bands", "no bands.", "none banded", "no banded birds",
    "no bands or flags", "no bands or flags.", "not banded", "not banded.",
    "none banded.", "no bands seen", "none visible", "all unbanded",
    "both were unbanded.", "both unbanded", "unbanded",
    "(not banded)", "no band", "no flag",
    # XX variants: all-unbanded shorthand
    "xx", "x:x", "x//x:x//x", "all xx", "all x:x", "2 xx", "3 xx",
    "4 xx", "5 xx", "6 xx", "7 xx", "8 xx", "9 xx", "10 xx",
}

def _is_no_band(text: str) -> bool:
    t = text.lower().strip().rstrip(".")
    if t in _NONE_BARE:
        return True
    # Starts-with variants
    if any(t.startswith(p) for p in (
        "no band", "not band", "none band", "no flag",
        "both were unbanded", "all unbanded", "all xx", "all x:x",
    )):
        return True
    # "N xx" or "N x:x" (e.g. "2 XX", "3 x:x")
    if re.fullmatch(r"\d+\s+(?:xx|x:x|unbanded)", t, re.IGNORECASE):
        return True
    return False


# ── Unreadable / count-unknown detection ─────────────────────────────────────
_UNREADABLE_PHRASES = [
    "none readable", "not readable", "unreadable",
    "too distant", "unable to read", "could not read",
    "partial resight", "incomplete band combo",
    "code not readable",
    "banded pipl.*reported directly to banders",  # "7 banded PIPL all reported..."
    "banded.*all reported",
]
_UNREADABLE_RE = re.compile(
    "|".join(_UNREADABLE_PHRASES), re.IGNORECASE
)

def _is_unreadable(text: str) -> bool:
    return bool(_UNREADABLE_RE.search(text))


# ── Trailing metadata noise ───────────────────────────────────────────────────
_TRAILING_NOISE = re.compile(
    r"[\-–]?\s*"
    r"(?:reported(?: to banders?| to (?:great lakes|audubon|vt|plover@\S+)"
    r"(?:[^.]*)?)?|"
    r"REPORTED TO BANDERS?|"
    r"not reported to another site|"
    r"will be batch reported[^.]*|"
    r"photos? (?:taken|available[^.]*)|"
    r"photo;[^)]*|"
    r"banded as (?:a chick|an adult)[^.]*|"
    r"formerly observed[^.]*|"
    r"this bird has been sighted[^.]*|"
    r"turned into banded birds[^.]*|"
    r"continuing bird[^.]*)"
    r"[.,]?\s*$",
    re.IGNORECASE,
)

# Leading "N banded (PIPL) (seen|birds|:)" prefix on the whole cell
_LEADING_COUNT = re.compile(
    r"^\d+\s+(?:banded\s+)?(?:pipl\s+)?(?:banded\s+)?(?:birds?\s+)?(?:banded\s+)?(?:seen[,.]?\s*|:?\s*)",
    re.IGNORECASE,
)

def _clean(entry: str) -> str:
    return _TRAILING_NOISE.sub("", entry).strip().strip(",").strip()

def _fmt(entries: list[str]) -> str:
    cleaned = [_clean(e) for e in entries if e and e.strip()]
    cleaned = [e for e in cleaned if e]
    return "\n".join(f"{i+1}) {e}" for i, e in enumerate(cleaned))


# ── Splitters ─────────────────────────────────────────────────────────────────

def _split_depth0(text: str, seps=(",", ";")) -> list[str]:
    """Split on ', ' or '; ' only when not inside parentheses."""
    parts, buf, depth = [], [], 0
    i = 0
    while i < len(text):
        c = text[i]
        if c == "(":
            depth += 1; buf.append(c)
        elif c == ")":
            depth -= 1; buf.append(c)
        elif depth == 0 and c in seps and i + 1 < len(text) and text[i + 1] == " ":
            parts.append("".join(buf).strip())
            buf = []; i += 2; continue
        else:
            buf.append(c)
        i += 1
    if buf:
        parts.append("".join(buf).strip())
    return [p for p in parts if p]


# Patterns where the observer already numbered birds
_NUMBERED_RE = [
    re.compile(r"(?:^|\s)(\d+)\)\s*"),                    # 1) 2) 3)
    re.compile(r"(?:^|\n)\s*(\d+)\.\s+"),                 # 1. 2. 3.
    re.compile(r"PP(\d+):\s*", re.IGNORECASE),            # PP1: PP2:
    re.compile(r"PIPL\s+([A-Z]):\s*", re.IGNORECASE),    # PIPL A: PIPL B:
    re.compile(r"\((\d+)\)\s*[-–]?\s*"),                  # (1) - ... (2) - ...
]

def _try_numbered(text: str) -> list[str] | None:
    for pat in _NUMBERED_RE:
        if pat.search(text):
            parts = re.split(pat.pattern, text, flags=pat.flags)
            entries = [p.strip() for p in parts
                       if p and p.strip() and not re.fullmatch(r"[\dA-Z]", p.strip())]
            if len(entries) >= 1:
                return entries
    return None


def _try_newlines(text: str) -> list[str] | None:
    """Each newline-separated line = one bird."""
    if "\n" not in text:
        return None
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    return lines if len(lines) >= 2 else None


def _try_slash_slash(text: str) -> list[str]:
    """Standard // notation — depth-0 comma/semicolon split."""
    body = re.sub(r"^\d+:\s*", "", text)      # strip "N: " count prefix
    return _split_depth0(body)


def _try_and_split(text: str) -> list[str] | None:
    """Split on ' and ' as a bird separator."""
    if " and " not in text.lower():
        return None
    parts = re.split(r"\s+and\s+", text, flags=re.IGNORECASE)
    # Filter out "X unbanded" trailing notes
    parts = [p.strip() for p in parts
             if p.strip() and not re.fullmatch(r"\d+\s+unbanded", p.strip(), re.IGNORECASE)]
    return parts if len(parts) >= 2 else None


def _try_ampersand(text: str) -> list[str] | None:
    """Split on ' & ' as a bird separator."""
    if "&" not in text:
        return None
    parts = [p.strip() for p in text.split("&") if p.strip()]
    return parts if len(parts) >= 2 else None


def _try_multispaces(text: str) -> list[str] | None:
    """
    Some observers separate multiple birds with 2+ spaces (no comma/newline).
    Only use when we get ≥2 parts that each look like a band entry.
    """
    if not re.search(r"  +", text):
        return None
    parts = [p.strip() for p in re.split(r"  +", text) if p.strip()]
    if len(parts) < 2:
        return None
    def looks_band_like(p):
        return bool(re.search(r"[A-Za-z][./:,\-\\]|//|\bF[OGBYKW]\b|\bS\b", p))
    if sum(looks_band_like(p) for p in parts) >= 2:
        return parts
    return None


def _try_semicolons(text: str) -> list[str] | None:
    """Semicolon-separated birds (without // notation)."""
    if ";" not in text:
        return None
    parts = [p.strip() for p in text.split(";") if p.strip()]
    if len(parts) < 2:
        return None
    def looks_band_like(p):
        p = p.lower()
        return any(k in p for k in (":", ",", "left", "right", "ll=", "rl=", "//", "flag", "band"))
    if sum(looks_band_like(p) for p in parts) >= 2:
        return parts
    return None


# ── Main parser ───────────────────────────────────────────────────────────────

def parse_band_cell(raw) -> tuple[str | None, bool]:
    """
    Returns (parsed_text, needs_review).

    parsed_text = "1) ...\n2) ..." — one numbered entry per bird.
    needs_review = True (pink) only when count is genuinely unknowable.
    Default for anything unclear = "1) <raw text>", no pink.
    """
    if raw is None:
        return None, False
    text = str(raw).strip()
    if not text or text == "-":
        return None, False

    # 1. No banded birds
    if _is_no_band(text):
        return "No banded birds", False

    # 2. Count genuinely unknowable → pink
    if _is_unreadable(text):
        return text, True

    # Strip leading "N banded (PIPL) seen/:" prefix
    text = _LEADING_COUNT.sub("", text).strip()
    if not text:
        return "No banded birds", False

    # 3. Observer already numbered birds (1), 1., PP1:, PIPL A:, (1) -)
    numbered = _try_numbered(text)
    if numbered:
        return _fmt(numbered), False

    # 4. Newline-separated — each line = one bird
    newlines = _try_newlines(text)
    if newlines:
        return _fmt(newlines), False

    # 5. Standard // notation → depth-0 comma/semicolon split
    if "//" in text:
        birds = _try_slash_slash(text)
        if birds:
            return _fmt(birds), False

    # 6. ' and ' as bird separator
    and_parts = _try_and_split(text)
    if and_parts:
        return _fmt(and_parts), False

    # 7. '&' as bird separator
    amp_parts = _try_ampersand(text)
    if amp_parts:
        return _fmt(amp_parts), False

    # 8. Multi-space separator (e.g. "FO.YY-S.G   GF.KO-S.B")
    ms_parts = _try_multispaces(text)
    if ms_parts:
        return _fmt(ms_parts), False

    # 9. Semicolon-separated without //
    semi_parts = _try_semicolons(text)
    if semi_parts:
        return _fmt(semi_parts), False

    # 10. Default: 1 bird, keep raw text
    return f"1) {_clean(text)}", False


# ── Find band columns ─────────────────────────────────────────────────────────

def find_band_cols(headers: list, band_col_re) -> list[int]:
    return [i for i, h in enumerate(headers) if h and band_col_re.search(str(h))]


# ── Process one year ──────────────────────────────────────────────────────────

def process_year(year: int):
    cfg      = YEAR_CONFIG[year]
    in_path  = CLEAN_IN / f"Winter Birds {year} Clean.xlsx"
    out_path = REVIEW   / f"Winter Birds {year} Band Review.xlsx"

    print(f"\n{'='*60}")
    print(f"  {year}  —  {in_path.name}")

    wb_in = openpyxl.load_workbook(in_path)
    ws_in = wb_in[cfg["sheet"]]

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = cfg["sheet"]

    headers   = [cell.value for cell in ws_in[1]]
    band_cols = set(find_band_cols(headers, cfg["band_col_re"]))

    print(f"  Sheet: '{cfg['sheet']}'  |  "
          f"Rows: {ws_in.max_row - 1}  |  Band cols: {len(band_cols)}")

    stats = {"total": 0, "no_band": 0, "parsed": 0, "pink": 0}

    for row_idx, row in enumerate(ws_in.iter_rows(), start=1):
        out_row = []
        results = {}

        for col_idx, cell in enumerate(row):
            if row_idx == 1 or col_idx not in band_cols:
                out_row.append(cell.value)
            else:
                parsed, needs_review = parse_band_cell(cell.value)
                results[col_idx] = (parsed, needs_review)
                out_row.append(parsed)

        ws_out.append(out_row)

        if row_idx > 1:
            for col_idx, (parsed, needs_review) in results.items():
                raw_val = ws_in.cell(row=row_idx, column=col_idx + 1).value
                if raw_val is not None and str(raw_val).strip() and str(raw_val).strip() != "-":
                    stats["total"] += 1
                    if parsed == "No banded birds":
                        stats["no_band"] += 1
                    elif needs_review:
                        stats["pink"] += 1
                    else:
                        stats["parsed"] += 1
                if needs_review:
                    ws_out.cell(row=row_idx, column=col_idx + 1).fill = PINK

    ws_out.freeze_panes = "A2"
    wb_out.save(out_path)

    print(f"  Saved: {out_path.name}")
    print(f"  Non-empty: {stats['total']}  |  "
          f"No band: {stats['no_band']}  |  "
          f"Numbered: {stats['parsed']}  |  "
          f"Pink: {stats['pink']}")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) > 1:
        years = [int(y) for y in sys.argv[1:]]
    else:
        years = list(YEAR_CONFIG.keys())

    for year in years:
        process_year(year)

    print("\nDone.")
