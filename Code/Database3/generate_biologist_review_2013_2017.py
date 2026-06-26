"""
generate_biologist_review_2013_2017.py
───────────────────────────────────────
Generates a biologist review workbook for Database 3 band data from the
2013–2017 winter surveys. Mirrors the format of the existing 2018–2024
workbook (Biologist_Band_Review_2018-2024.xlsx) so the biologist sees the
same 8-column review layout they already know.

For each row with non-trivial PIPL band info:
  Route | Point | Date | PIPL@Pt | Original Text | Our Interpretation | Our Notes | Your Correction

Three transformations applied on top of the existing interpret() logic:
  1. Non-PIPL species (SNPL/REKN/WIPL/AMOY/DCCO/DUNL/etc.) are stripped from
     band text, since 2013/2014 often have multi-species data crammed into
     the band cell. A note ("Stripped N non-PIPL entries") is recorded so
     the biologist can audit our species filtering.
  2. Multi-line "Upper left: X / Lower left: Y / Upper right: Z / Lower right: W"
     descriptions (very common in 2014) are compacted to a single-line
     "UL: X, LL: Y, UR: Z, LR: W" format that matches the 2015+ style.
  3. PIPL count at each point is parsed from messy free-text species fields
     ("PIPL- 8, SNPL- 4" → 8; "PIPL (5 TOTAL, 3 banded)" → 5).

Output: Databases/Database3BiologistReview/Biologist_Band_Review_2013-2017.xlsx
"""

import re
import sys
import os
from pathlib import Path
from datetime import datetime, time as datetime_time

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Reuse helpers from the 2018–2024 generator ───────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from generate_biologist_review import (
    interpret,
    is_no_band, is_unreadable,
    C_HEADER, C_OK, C_REVIEW, C_CONFLICT, C_INST_HDR, C_INST_BG,
)


ROOT    = Path(__file__).resolve().parents[2]
CLEAN   = ROOT / "Databases" / "Database3Clean"
OUT_DIR = ROOT / "Databases" / "Database3BiologistReview"
OUT_DIR.mkdir(exist_ok=True)
OUT     = OUT_DIR / "Biologist_Band_Review_2013-2017.xlsx"


# ══════════════════════════════════════════════════════════════════════════════
# Non-PIPL species stripping
# ══════════════════════════════════════════════════════════════════════════════

# Common non-PIPL species codes that appear in band/species fields.
# Anything matching one of these followed by counts/colons/parens gets removed
# from band text before interpretation.
NON_PIPL_CODES = [
    "SNPL", "REKN", "WIPL", "AMOY",
    "DCCO", "DUNL", "RUTU", "SAND", "SBDO", "BBPL", "SEPL", "WESA", "LESA",
    "WILL", "REEG", "SPPL", "GBHE", "BCNH", "LAGU", "RBGU",
    # Spelled-out versions occasionally appear too
    "Snowy Plover", "Red Knot", "Wilson's Plover", "American Oystercatcher",
    "Black-bellied Plover", "Semi-palmated Plover", "Dunlin", "Sanderling",
    "Ruddy Turnstone", "Western Sandpiper", "Least Sandpiper",
    "Short-billed Dowitcher", "Willet", "Double-crested Cormorant",
]


def _strip_non_pipl_chunk(text: str) -> tuple[str, int]:
    """
    Remove segments that look like "<N> banded SNPL: ..." or "SNPL (3)" etc.
    Returns (cleaned_text, n_stripped) where n_stripped is the number of
    distinct non-PIPL species segments removed.
    """
    if not text:
        return text, 0

    n_stripped = 0
    out = text

    # Pattern A: "<N> banded SNPL: ..., ..." up to next species marker / end
    for code in NON_PIPL_CODES:
        # match e.g. "15 banded SNPL: WK:S//, S//:KW, OY:S//W" up to next ", N banded XXXX" or " N banded XXXX" or end
        pat = re.compile(
            rf"\b\d*\s*banded\s+{re.escape(code)}\s*:?\s*[^.]*?"
            rf"(?=(?:[,;.\s]+\d+\s+banded\s+[A-Z][a-zA-Z]{{2,}}|$))",
            re.IGNORECASE,
        )
        new = pat.sub("", out)
        if new != out:
            n_stripped += 1
            out = new

    # Pattern B: bare code with count "SNPL (3)" or "SNPL-3" or "SNPL 3"
    for code in NON_PIPL_CODES:
        pat = re.compile(rf"\b{re.escape(code)}\b\s*[-:]?\s*\(?\s*\d*\s*\)?", re.IGNORECASE)
        new = pat.sub("", out)
        if new != out and code.upper() != "PIPL":
            n_stripped += 1
            out = new

    # Pattern C: "Piping Plover:" / "PIPL:" — leading label that we can drop
    out = re.sub(r"^\s*(?:Piping\s+Plover|PIPL)\s*:?\s*", "", out, flags=re.IGNORECASE)

    # Cleanup: collapse multiple spaces, dangling punctuation
    out = re.sub(r"\s{2,}", " ", out).strip(" ,;.\n\t")

    return out, n_stripped


# ══════════════════════════════════════════════════════════════════════════════
# Multi-line "Upper left: X / Lower left: Y / …" compactor
# ══════════════════════════════════════════════════════════════════════════════

# Maps various spellings of leg positions to compact tokens.
_LEG_MAP = [
    (re.compile(r"upper\s*left",  re.IGNORECASE), "UL"),
    (re.compile(r"lower\s*left",  re.IGNORECASE), "LL"),
    (re.compile(r"upper\s*right", re.IGNORECASE), "UR"),
    (re.compile(r"lower\s*right", re.IGNORECASE), "LR"),
]

def _compact_leg_description(text: str) -> str:
    """
    Convert:
        Upper left: FO
        Lower left: Y/Lb
        Upper right: -
        Lower right: Y
    into:
        UL: FO, LL: Y/Lb, UR: -, LR: Y
    Operates per-line and only if multiple legs are described.
    """
    if not text or "\n" not in text:
        return text
    # Compact each leg term
    out = text
    for pat, token in _LEG_MAP:
        out = pat.sub(token, out)
    # Collapse newlines between leg fragments into ", "
    lines = [l.strip().rstrip(",").rstrip(";") for l in out.split("\n") if l.strip()]
    # Only compact if at least 2 lines start with a compact token (UL/LL/UR/LR)
    leg_lines = [l for l in lines if re.match(r"^(UL|LL|UR|LR)\s*[:\-]", l, re.IGNORECASE)]
    if len(leg_lines) >= 2:
        # Keep any non-leg leading line as a prefix (e.g. "PIPL")
        prefix_lines = [l for l in lines if not re.match(r"^(UL|LL|UR|LR)\s*[:\-]", l, re.IGNORECASE)]
        prefix = ", ".join(prefix_lines) + ": " if prefix_lines else ""
        # Strip the "PIPL" prefix specifically — we know it's PIPL data
        prefix = re.sub(r"^PIPL\s*[:\-,]?\s*", "", prefix, flags=re.IGNORECASE)
        return prefix + ", ".join(leg_lines)
    return text


# ══════════════════════════════════════════════════════════════════════════════
# Parse PIPL count from messy species text
# ══════════════════════════════════════════════════════════════════════════════

_PIPL_COUNT_PATTERNS = [
    re.compile(r"PIPL\s*[-:]?\s*\(?\s*(\d+)\b",            re.IGNORECASE),   # "PIPL- 8", "PIPL: 5", "PIPL (5)"
    re.compile(r"(\d+)\s*PIPL\b",                          re.IGNORECASE),   # "46 PIPL"
    re.compile(r"PIPL\s*\(\s*(\d+)\b",                     re.IGNORECASE),   # "PIPL (1)"
    re.compile(r"Piping\s*Plover\s*[-:]?\s*\(?\s*(\d+)\b", re.IGNORECASE),
]

def parse_pipl_count(text) -> int:
    """Extract the PIPL count from a free-text species description.
    Returns 0 if none found."""
    if text is None:
        return 0
    s = str(text)
    for pat in _PIPL_COUNT_PATTERNS:
        m = pat.search(s)
        if m:
            try:
                return int(m.group(1))
            except ValueError:
                continue
    # Fallback: if "PIPL" appears with no number, assume 1
    if re.search(r"\bPIPL\b|\bPiping\s*Plover\b", s, re.IGNORECASE):
        return 1
    return 0


# ══════════════════════════════════════════════════════════════════════════════
# Extended interpreter: strip non-PIPL → compact legs → existing interpret()
# ══════════════════════════════════════════════════════════════════════════════

def interpret_extended(raw, pipl_count: int) -> tuple[list, str, str]:
    """
    Pipeline:
      1. Strip non-PIPL species segments
      2. Compact multi-line leg descriptions
      3. Run the shared interpret() from generate_biologist_review.py
    Returns (entries, status, notes) — same shape as the original interpret().
    """
    if raw is None or str(raw).strip() == "":
        return [], "ok", ""

    text = str(raw).strip()
    extra_notes = []

    # Stage 1: strip non-PIPL species
    cleaned, n_stripped = _strip_non_pipl_chunk(text)
    if n_stripped > 0:
        extra_notes.append(
            f"Stripped {n_stripped} non-PIPL species segment(s) — please confirm only PIPL data remains"
        )

    # If everything was non-PIPL, skip
    if not cleaned or is_no_band(cleaned):
        if extra_notes:
            return [], "review", "  |  ".join(extra_notes)
        return [], "ok", ""

    # Stage 2: compact multi-line leg descriptions
    cleaned = _compact_leg_description(cleaned)

    # Stage 3: existing interpret()
    entries, status, notes = interpret(cleaned, pipl_count)

    if extra_notes:
        combined_notes = "  |  ".join(extra_notes + ([notes] if notes else []))
        # If we stripped non-PIPL, always mark for review
        if status == "ok":
            status = "review"
        return entries, status, combined_notes
    return entries, status, notes


# ══════════════════════════════════════════════════════════════════════════════
# Date / value helpers
# ══════════════════════════════════════════════════════════════════════════════

def _format_date(dv) -> str:
    if dv is None:
        return ""
    if hasattr(dv, "strftime"):
        try:
            return dv.strftime("%Y-%m-%d")
        except Exception:
            pass
    return str(dv)[:10]


def _has_band_content(*cells) -> bool:
    """True if any of the given cells contain non-trivial band content."""
    for c in cells:
        if c is None:
            continue
        s = str(c).strip()
        if not s or s.lower() in ("0", "none", "n/a", "na", "-"):
            continue
        if is_no_band(s):
            continue
        return True
    return False


# ══════════════════════════════════════════════════════════════════════════════
# Per-year readers
# ══════════════════════════════════════════════════════════════════════════════

def read_2013():
    """
    2013 — Species GPS sheet
      Row 1: annotations (skip)
      Row 2: headers
      Row 3+: data
    Band info is in col 9 (Flag color/code) + col 10 (Color band combo).
    PIPL count comes from col 2 "Focal species" text.
    """
    wb = openpyxl.load_workbook(CLEAN / "Winter Birds '13 Clean.xlsx", read_only=True)
    ws = wb["Species GPS"]

    records = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        date_val = row[0] if len(row) > 0 else None
        route    = row[1] if len(row) > 1 else None
        species  = row[2] if len(row) > 2 else None
        flag     = row[9] if len(row) > 9 else None
        combo    = row[10] if len(row) > 10 else None

        if not route:
            continue

        pipl = parse_pipl_count(species)
        if pipl == 0:
            continue   # no PIPL at this location → not a banding row

        # Combine flag + combo columns. Treat literal "0" as null.
        parts = []
        for c in (flag, combo):
            if c is None:
                continue
            s = str(c).strip()
            if s and s != "0" and s.lower() not in ("none", "n/a", "na"):
                parts.append(s)
        if not parts:
            continue
        raw = " | ".join(parts) if len(parts) > 1 else parts[0]

        if is_no_band(raw):
            continue

        entries, status, notes = interpret_extended(raw, pipl)
        if not entries and status == "ok":
            continue

        records.append({
            "route":   str(route).strip(),
            "point":   "",      # 2013 Species GPS has no group/point number
            "date":    _format_date(date_val),
            "pipl":    pipl,
            "raw":     raw,
            "entries": entries,
            "status":  status,
            "notes":   notes,
        })

    wb.close()
    return records


def read_2014():
    """
    2014 — Indiv Flock GPS & bands sheet
      Row 1: instructions (skip)
      Row 2: empty/header continuation (skip)
      Row 3: real headers
      Row 4+: data
    Band info is in col 7 (one column, often messy with multi-species).
    PIPL count from col 6 "Species and number of individuals".
    """
    wb = openpyxl.load_workbook(CLEAN / "Winter Birds '14 Clean.xlsx", read_only=True)
    ws = wb["Indiv Flock GPS & bands"]

    records = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        date_val = row[0] if len(row) > 0 else None
        route    = row[1] if len(row) > 1 else None
        point    = row[3] if len(row) > 3 else None
        species  = row[6] if len(row) > 6 else None
        bands    = row[7] if len(row) > 7 else None

        if not route:
            continue

        pipl = parse_pipl_count(species)
        if pipl == 0:
            continue

        if bands is None or str(bands).strip() == "":
            continue
        raw = str(bands).strip()
        if is_no_band(raw):
            continue

        entries, status, notes = interpret_extended(raw, pipl)
        if not entries and status == "ok":
            continue

        records.append({
            "route":   str(route).strip(),
            "point":   point if point is not None else "",
            "date":    _format_date(date_val),
            "pipl":    pipl,
            "raw":     raw,
            "entries": entries,
            "status":  status,
            "notes":   notes,
        })

    wb.close()
    return records


def _build_pipl_lookup_for_year(wb, hdr_row: int) -> dict:
    """
    For 2015–2017, build a (route, date, point) → PIPL count lookup
    by parsing the "Species and number of individuals" column on DATA SHEET 1.
    """
    ws = wb["DATA SHEET 1"]
    lookup = {}
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if len(row) < 15:
            continue
        date_val = row[0]
        route    = row[3]
        point    = row[11]
        species  = row[14]
        if not route or species is None:
            continue
        pipl = parse_pipl_count(species)
        if pipl == 0:
            continue
        key = (str(route).strip().lower(), _format_date(date_val), str(point).strip() if point is not None else "")
        # First occurrence wins
        if key not in lookup:
            lookup[key] = pipl
    return lookup


def _compose_2015_2017_combo(flag_code, flag_col, orient, ul, ll, ur, lr) -> str:
    """
    For 2015-2017, each source row = ONE banded bird.
    Compose the bands into a single descriptive string we can show as both
    the Original Text and the Our Interpretation. NOT split into multiple
    "birds" — these are leg positions of ONE bird.
    """
    parts = []
    if flag_code and str(flag_code).strip() and str(flag_code).strip() != "-":
        parts.append(f"Code: {str(flag_code).strip()}")
    if flag_col and str(flag_col).strip() and str(flag_col).strip() != "-":
        parts.append(f"Color: {str(flag_col).strip()}")
    if orient and str(orient).strip() and str(orient).strip() != "-":
        parts.append(f"Orientation: {str(orient).strip()}")
    for label, val in (("UL", ul), ("LL", ll), ("UR", ur), ("LR", lr)):
        if val is not None and str(val).strip():
            parts.append(f"{label}: {str(val).strip()}")
    return "; ".join(parts)


def read_2015_2017(year: int):
    """
    2015 / 2016 / 2017 — DATA SHEET 3
      EACH ROW = ONE BANDED PIPL BIRD at a (route, date, point).
      Columns: 7=Band/Flag Code, 8=Color, 9=Orientation,
               10=UL, 11=LL, 12=UR, 13=LR
    PIPL count looked up from DATA SHEET 1 col 14.

    Because the source already separates birds row-by-row, we DO NOT
    pass these through the multi-bird splitter. Each review row holds
    one bird's full description as a single entry.

    To detect count conflicts at the (route, date, point) level, we
    aggregate the count of banded PIPL rows seen and compare to PIPL@Pt
    in a second pass.
    """
    fname = {
        2015: "Winter Birds '15 Clean.xlsx",
        2016: "Winter Birds '16 Clean.xlsx",
        2017: "Winter Birds '17 Clean.xlsx",
    }[year]
    wb = openpyxl.load_workbook(CLEAN / fname, read_only=True)

    hdr_row = 1 if year == 2017 else 2
    pipl_lookup = _build_pipl_lookup_for_year(wb, hdr_row)

    ws = wb["DATA SHEET 3"]
    records = []
    # Track how many banded entries we've collected per (route, date, point)
    per_point_band_count: dict = {}

    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if len(row) < 14:
            continue
        date_val = row[0]
        route    = row[3]
        point    = row[5]
        species  = row[6]
        flag_code = row[7]
        flag_col  = row[8]
        orient    = row[9]
        ul, ll, ur, lr = row[10], row[11], row[12], row[13]

        if not route or not species:
            continue
        sp = str(species).strip().upper()
        if sp not in ("PIPL", "PIPING PLOVER"):
            continue
        if not _has_band_content(flag_code, flag_col, ul, ll, ur, lr):
            continue

        raw = _compose_2015_2017_combo(flag_code, flag_col, orient, ul, ll, ur, lr)
        if not raw:
            continue

        key = (str(route).strip().lower(), _format_date(date_val),
               str(point).strip() if point is not None else "")
        pipl = pipl_lookup.get(key, 0)
        per_point_band_count[key] = per_point_band_count.get(key, 0) + 1

        # Single-entry record — each source row is one bird, no splitting
        records.append({
            "route":   str(route).strip(),
            "point":   point if point is not None else "",
            "date":    _format_date(date_val),
            "pipl":    pipl,
            "raw":     raw,
            "entries": [raw],
            "status":  "ok",
            "notes":   "",
            "_key":    key,
        })

    wb.close()

    # Second pass: detect (route, date, point) groups where total banded > PIPL@Pt
    conflicts = {k for k, n in per_point_band_count.items()
                 if pipl_lookup.get(k, 0) > 0 and n > pipl_lookup[k]}
    missing_pipl_lookup = {k for k in per_point_band_count
                           if k not in pipl_lookup}

    for rec in records:
        k = rec.pop("_key")
        notes = []
        if k in conflicts:
            n_banded = per_point_band_count[k]
            notes.append(
                f"⚠ CONFLICT: {n_banded} banded PIPL rows at this point but "
                f"only {rec['pipl']} PIPL counted. Please correct."
            )
            rec["status"] = "conflict"
        elif k in missing_pipl_lookup:
            notes.append(
                "PIPL count at this point could not be looked up from DATA SHEET 1 — "
                "please verify."
            )
            rec["status"] = "review"
        if notes:
            rec["notes"] = "  |  ".join(notes)

    return records


# ══════════════════════════════════════════════════════════════════════════════
# Workbook builder (same shape as the 2018–2024 generator's)
# ══════════════════════════════════════════════════════════════════════════════

def make_workbook(records_by_year: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Instructions sheet ────────────────────────────────────────────────────
    ws_inst = wb.create_sheet("How to Fill This")
    ws_inst.sheet_view.showGridLines = False

    inst_header_fill = PatternFill("solid", fgColor=C_INST_HDR)
    inst_body_fill   = PatternFill("solid", fgColor=C_INST_BG)
    white_fill       = PatternFill("solid", fgColor="FFFFFF")
    bold_white       = Font(bold=True, color="FFFFFF", size=13)
    normal           = Font(size=11)

    def inst_row(ws, row_num, text, is_header=False, is_blank=False):
        cell = ws.cell(row=row_num, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if is_header:
            cell.fill = inst_header_fill
            cell.font = bold_white
        elif is_blank:
            cell.fill = white_fill
        else:
            cell.fill = inst_body_fill
            cell.font = normal
        ws.row_dimensions[row_num].height = 30 if is_header else 22

    ws_inst.column_dimensions["A"].width = 3
    ws_inst.column_dimensions["B"].width = 90

    content = [
        ("USFWS Piping Plover Winter Survey — Band Data Review (2013–2017)", True),
        ("", False),
        ("Thank you for reviewing this file. We have processed the raw banding notes "
         "from the 2013–2017 winter surveys. These early years recorded band data in "
         "less structured formats than later years, so your expertise is especially "
         "valuable here. Each year is a separate sheet.", False),
        ("", False),
        ("WHAT WE CLEANED BEFORE YOU SEE THE DATA", True),
        ("", False),
        ("• Non-PIPL species (SNPL, REKN, WIPL, AMOY, etc.) were stripped from band "
         "text. The original row text is shown in Column E so you can confirm we "
         "didn't accidentally remove PIPL data.", False),
        ("• Multi-line leg descriptions (Upper left: X / Lower left: Y / ...) were "
         "compacted to single-line format (UL: X, LL: Y, UR: Z, LR: W) to match "
         "the 2015+ style.", False),
        ("• Literal placeholders like '0', 'none', 'n/a' in band cells were treated "
         "as 'no banded birds' and excluded from review.", False),
        ("", False),
        ("HOW TO FILL THIS IN — please read carefully", True),
        ("", False),
        ("Each row is one banding cell from the original data sheet "
         "(one route + one point + one survey).", False),
        ("", False),
        ("Column E  'Original Text'      — Exactly what was in the field data. Read-only.", False),
        ("Column F  'Our Interpretation' — How we split it into individual birds. Read-only.", False),
        ("Column G  'Our Notes'          — Any flag or issue we found. Read-only.", False),
        ("", False),
        ("Column H  'Your Correction'    — THE ONLY COLUMN YOU NEED TO FILL IN.", False),
        ("", False),
        ("  ✓  If Column F looks correct  →  leave Column H BLANK. No action needed.", False),
        ("", False),
        ("  ✗  If Column F is wrong  →  write the correct version in Column H\n"
         "     using numbered bullets, one bird per line, like this:\n"
         "\n"
         "         1) Of//GG:S//K\n"
         "         2) Gf//OO:X//GB\n"
         "         3) Yf(O64)//WK:S//bW\n"
         "\n"
         "     Each line = one individual bird's band combination.\n"
         "     If there is only 1 bird, just write:  1) <band combo>", False),
        ("", False),
        ("COLOUR GUIDE", True),
        ("", False),
        ("  White rows  — We are confident. A quick glance is enough.", False),
        ("  Yellow rows — We are unsure about the count, split, or species stripping. "
         "Please check carefully.", False),
        ("  Orange rows — Count conflict (more banded entries than total PIPL at that "
         "point). These definitely need your correction.", False),
        ("", False),
        ("ONE MORE THING", True),
        ("", False),
        ("Column D 'PIPL at Point' is the total Piping Plovers counted at that location.\n"
         "The number of banded birds in Column H cannot exceed this number.\n"
         "For example: if PIPL at Point = 3, you can have at most 3 banded entries.", False),
        ("", False),
        ("Thank you — please email the completed file back when done.", False),
    ]
    for i, (text, is_hdr) in enumerate(content, start=2):
        inst_row(ws_inst, i, text, is_header=is_hdr, is_blank=(text == "" and not is_hdr))

    # ── Per-year sheets ───────────────────────────────────────────────────────
    STATUS_FILL = {
        "ok":       PatternFill("solid", fgColor=C_OK),
        "review":   PatternFill("solid", fgColor=C_REVIEW),
        "conflict": PatternFill("solid", fgColor=C_CONFLICT),
    }
    HDR_COLS = ["Route", "Point", "Date", "PIPL at Point",
                "Original Text", "Our Interpretation",
                "Our Notes (read only)", "Your Correction — leave blank if Col F is correct"]
    COL_WIDTHS = [38, 7, 12, 14, 50, 50, 38, 45]

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor=C_HEADER)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(border_style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for year, records in records_by_year.items():
        ws = wb.create_sheet(str(year))
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        for ci, (col_name, width) in enumerate(zip(HDR_COLS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = cell_border
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 28

        for ri, rec in enumerate(records, start=2):
            interpretation = "\n".join(
                f"{i+1}) {e}" for i, e in enumerate(rec["entries"])
            ) if rec["entries"] else "(no banded birds — all unbanded)"

            values = [
                rec["route"], rec["point"], rec["date"], rec["pipl"],
                rec["raw"], interpretation, rec["notes"], "",
            ]
            fill = STATUS_FILL[rec["status"]]
            for ci, val in enumerate(values, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill      = fill
                cell.border    = cell_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            n_lines = max(
                len(str(rec["raw"]).split("\n")),
                len(rec["entries"]) + 1,
                1,
            )
            ws.row_dimensions[ri].height = max(18, n_lines * 16)

    return wb


# ══════════════════════════════════════════════════════════════════════════════
# Run
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    readers = {
        2013: read_2013,
        2014: read_2014,
        2015: lambda: read_2015_2017(2015),
        2016: lambda: read_2015_2017(2016),
        2017: lambda: read_2015_2017(2017),
    }

    records_by_year = {}
    total_ok = total_review = total_conflict = 0

    for year, reader in readers.items():
        print(f"Reading {year}...")
        recs = reader()
        records_by_year[str(year)] = recs
        ok       = sum(1 for r in recs if r["status"] == "ok")
        review   = sum(1 for r in recs if r["status"] == "review")
        conflict = sum(1 for r in recs if r["status"] == "conflict")
        total_ok += ok; total_review += review; total_conflict += conflict
        print(f"  {len(recs):3d} cells  —  "
              f"White: {ok}  Yellow: {review}  Orange: {conflict}")

    print(f"\nAll years total — White: {total_ok}  "
          f"Yellow: {total_review}  Orange: {total_conflict}")

    print("\nBuilding Excel workbook...")
    wb = make_workbook(records_by_year)
    wb.save(OUT)
    print(f"Saved: {OUT}")
