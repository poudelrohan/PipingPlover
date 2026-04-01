"""
Step 6: Final Report Generation
─────────────────────────────────
Combines all pipeline outputs into a single 3-sheet Excel file:

  Sheet 1 — Clean_Data
  Sheet 2 — Removed_Rows
  Sheet 3 — Summary_Report

Output: db3_<year>_FINAL.xlsx  (in Output/<year>/)
"""

import pandas as pd
import os
import sys
from datetime import datetime

# ── Load config ────────────────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from database3_config import config, get_output_folder, get_filename

# ── Resolve paths ──────────────────────────────────────────────────────────────
script_dir    = os.path.dirname(os.path.abspath(__file__))
output_folder = get_output_folder(script_dir)

year     = config["active_year"]
year_cfg = config["years"][year]

clean_path   = os.path.join(output_folder, get_filename("clean"))
removed_path = os.path.join(output_folder, get_filename("removed"))
ids_path     = os.path.join(output_folder, get_filename("with_ids"))
final_path   = os.path.join(output_folder, get_filename("FINAL"))

# ── Load data ──────────────────────────────────────────────────────────────────
if not os.path.exists(clean_path):
    print(f"[ERROR] Missing file: {clean_path}")
    print("        Run all previous steps first.")
    sys.exit(1)

if not os.path.exists(ids_path):
    print(f"[ERROR] Missing file: {ids_path}")
    print("        Run all previous steps first.")
    sys.exit(1)

clean_df = pd.read_excel(clean_path)
original = pd.read_excel(ids_path)

if os.path.exists(removed_path):
    removed_df = pd.read_excel(removed_path)
else:
    removed_df = pd.DataFrame()

print(f"  Clean rows   : {len(clean_df)}")
print(f"  Removed rows : {len(removed_df)}")
print(f"  Total input  : {len(original)}")

# ── Extract internal columns before dropping from Clean_Data ──────────────────
geo_warnings = []
if "_geo_warning" in clean_df.columns:
    warned = clean_df[clean_df["_geo_warning"].notna()]
    for _, row in warned.iterrows():
        geo_warnings.append({
            "unique_id": row.get("unique_id"),
            "warning":   row["_geo_warning"],
        })
    clean_df = clean_df.drop(columns=["_geo_warning"])

geo_corrections = []
if "_geo_correction" in clean_df.columns:
    corrected = clean_df[clean_df["_geo_correction"].notna()]
    for _, row in corrected.iterrows():
        geo_corrections.append({
            "unique_id":  row.get("unique_id"),
            "correction": row["_geo_correction"],
        })
    clean_df = clean_df.drop(columns=["_geo_correction"])

# ── Sheet 2: Removed_Rows ──────────────────────────────────────────────────────
if len(removed_df) > 0:
    removed_sheet = removed_df.copy()
    removed_sheet = removed_sheet.rename(columns={"_removal_reason": "removal_reason"})

    for col in ["_geo_warning", "_geo_correction"]:
        if col in removed_sheet.columns:
            removed_sheet = removed_sheet.drop(columns=[col])

    front_cols = ["unique_id", "removal_reason", "source_database", "source_file", "source_sheet"]
    clean_cols = [c for c in clean_df.columns
                  if c not in front_cols and c in removed_sheet.columns]

    final_removed_cols = (
        [c for c in front_cols if c in removed_sheet.columns] +
        clean_cols
    )
    removed_sheet = removed_sheet[final_removed_cols]
else:
    removed_sheet = pd.DataFrame(columns=["unique_id", "removal_reason"])

# ── Sheet 3: Summary_Report ────────────────────────────────────────────────────
total_input   = len(original)
total_clean   = len(clean_df)
total_removed = len(removed_df)

if len(removed_df) > 0:
    geo_removals     = removed_df[removed_df["_removal_reason"].str.startswith("Outside Florida", na=False)]
    missing_removals = removed_df[removed_df["_removal_reason"].str.startswith("Missing", na=False)]
    dup_removals     = removed_df[removed_df["_removal_reason"].str.startswith("Duplicate", na=False)]
else:
    geo_removals = missing_removals = dup_removals = pd.DataFrame()

geo_reason_counts = geo_removals["_removal_reason"].value_counts().reset_index() if len(geo_removals) > 0 else pd.DataFrame()
if len(geo_reason_counts) > 0:
    geo_reason_counts.columns = ["removal_reason", "count"]

missing_reason_counts = missing_removals["_removal_reason"].value_counts().reset_index() if len(missing_removals) > 0 else pd.DataFrame()
if len(missing_reason_counts) > 0:
    missing_reason_counts.columns = ["removal_reason", "count"]

# Partial location warnings
loc_fields = config["location_fields"]["fields"]
partial_warnings = []
for _, row in clean_df.iterrows():
    missing = [f for f in loc_fields if f in clean_df.columns and (pd.isna(row.get(f)) or str(row.get(f)).strip() == "")]
    if 0 < len(missing) < len(loc_fields):
        partial_warnings.append({
            "unique_id":      row.get("unique_id"),
            "missing_fields": ", ".join(missing)
        })

# ── Build summary rows ───────────────────────────────────────────────────────
SECTION = "\u00a7"

summary_rows = [
    ("Run date",                                            datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("Database",                                            config["database_name"]),
    ("Source database label",                               config["source_database"]),
    ("Year processed",                                      year),
    (f"{SECTION}Input",                                     ""),
    ("Total rows input",                                    total_input),
    ("Source file",                                         year_cfg["file"]),
    ("Source sheet",                                        year_cfg["sheet"]),
    (f"{SECTION}Output",                                    ""),
    ("Total rows in Clean_Data",                            total_clean),
    ("Total rows removed",                                  total_removed),
    ("Percentage of rows kept",                             f"{round(total_clean / total_input * 100, 2) if total_input > 0 else 0}%"),
]

summary_rows.append((f"{SECTION}Geography Removals", ""))
summary_rows.append(("Total removed for geography", len(geo_removals)))
if len(geo_removals) > 0:
    for _, r in geo_reason_counts.iterrows():
        summary_rows.append((f"  {r['removal_reason']}", f"{r['count']} rows"))
else:
    summary_rows.append(("  None", ""))

summary_rows.append((f"{SECTION}Missing Field Removals", ""))
summary_rows.append(("Total removed for missing fields", len(missing_removals)))
if len(missing_removals) > 0:
    for _, r in missing_reason_counts.iterrows():
        summary_rows.append((f"  {r['removal_reason']}", f"{r['count']} rows"))
else:
    summary_rows.append(("  None", ""))

summary_rows.append((f"{SECTION}Duplicate Removals", ""))
summary_rows.append(("Total duplicates removed", len(dup_removals)))
summary_rows.append(("Criteria used", ", ".join(config["duplicate_criteria"])))

summary_rows.append((f"{SECTION}Coordinate Corrections", ""))
summary_rows.append(("Rows with auto-corrected coordinates", len(geo_corrections)))
if geo_corrections:
    for c in geo_corrections:
        summary_rows.append((f"  unique_id {c['unique_id']}", c["correction"]))
else:
    summary_rows.append(("  None", ""))

summary_rows.append((f"{SECTION}Location Warnings", ""))
summary_rows.append(("Rows with partial location data (kept but flagged)", len(partial_warnings)))
if partial_warnings:
    for w in partial_warnings:
        summary_rows.append((f"  unique_id {w['unique_id']}", f"Missing: {w['missing_fields']}"))
else:
    summary_rows.append(("  None", ""))

summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

# ── Write final Excel ─────────────────────────────────────────────────────────
with pd.ExcelWriter(final_path, engine="openpyxl") as writer:
    clean_df.to_excel(writer,      sheet_name="Clean_Data",     index=False)
    removed_sheet.to_excel(writer, sheet_name="Removed_Rows",   index=False)
    summary_df.to_excel(writer,    sheet_name="Summary_Report", index=False)

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = writer.book

    header_font  = Font(bold=True, color="FFFFFF")
    section_font = Font(bold=True, color="1565C0")
    header_fills = {
        "Clean_Data":     PatternFill("solid", fgColor="2E7D32"),
        "Removed_Rows":   PatternFill("solid", fgColor="C62828"),
        "Summary_Report": PatternFill("solid", fgColor="1565C0"),
    }

    date_cols = {"SurveyDate"}

    for sheet_name in ["Clean_Data", "Removed_Rows", "Summary_Report"]:
        ws   = wb[sheet_name]
        fill = header_fills[sheet_name]

        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")

        if sheet_name in ["Clean_Data", "Removed_Rows"]:
            for col in ws.iter_cols(min_row=1, max_row=1):
                if col[0].value in date_cols:
                    col_letter = get_column_letter(col[0].column)
                    for row in ws[col_letter]:
                        if row.row > 1:
                            row.number_format = "MM/DD/YYYY"

        if sheet_name == "Summary_Report":
            for row in ws.iter_rows(min_row=2):
                cell = row[0]
                if cell.value and str(cell.value).startswith(SECTION):
                    cell.value = str(cell.value).replace(SECTION, "")
                    for c in row:
                        c.font = section_font
                        c.fill = PatternFill("solid", fgColor="E3F2FD")

        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

print(f"\n[DONE] Step 6 complete ({year})")
print(f"  Final report : {final_path}")
print(f"  Sheets       : Clean_Data ({total_clean} rows) | Removed_Rows ({total_removed} rows) | Summary_Report")
