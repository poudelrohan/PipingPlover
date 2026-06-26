"""
Step 0: Extract PIPL Data — 2013
──────────────────────────────────
Reads Winter Birds '13 Clean.xlsx (Species GPS sheet) and emits one row
per individual bird (Option A expansion), applying biologist corrections
from Biologist_Band_Review_2013-2017.xlsx sheet "2013".

2013-specific handling:
  - Source has no Group/Point #. GroupNumber stays null.
  - PIPL count is extracted from the "Focal species" free-text column
    (e.g. "PIPL- 8, SNPL- 4, WIPL- 3" → 8).
  - Two band-bearing columns (col 9 Flag + col 10 Color band combo) are
    concatenated before parsing, with literal "0" treated as null.
  - Band info goes into BandCombo (NOT FlagCode/FlagColor — those stay null).

Output: db3_2013_extracted.xlsx in Output/2013/
"""

import pandas as pd
import re
import os
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
from database3_config import config, get_output_folder, get_filename

script_dir    = os.path.dirname(os.path.abspath(__file__))
output_folder = get_output_folder(script_dir)
year          = config["year"]

source_path = os.path.normpath(
    os.path.join(script_dir, config["input_folder"], config["file"])
)
review_path = (
    Path(script_dir).parents[2] / "Databases" / "Database3BiologistReview"
    / config["biologist_review_file"]
)
output_path = os.path.join(output_folder, get_filename("extracted"))

cols = config["columns"]


# ══════════════════════════════════════════════════════════════════════════════
# Helpers (shared across DB3 pipelines)
# ══════════════════════════════════════════════════════════════════════════════

def normalize_route(s):
    s = str(s or "").lower().strip()
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def _format_date(dv):
    if dv is None:
        return ""
    if hasattr(dv, "strftime"):
        try:
            return dv.strftime("%Y-%m-%d")
        except Exception:
            pass
    return str(dv)[:10]


def parse_coord(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        s = str(val).strip().rstrip("°").strip()
        f = float(s)
        return f if f != 0.0 else None
    except (ValueError, TypeError):
        return None


_PIPL_COUNT_PATTERNS = [
    re.compile(r"PIPL\s*[-:]?\s*\(?\s*(\d+)\b",            re.IGNORECASE),
    re.compile(r"(\d+)\s*PIPL\b",                          re.IGNORECASE),
    re.compile(r"PIPL\s*\(\s*(\d+)\b",                     re.IGNORECASE),
    re.compile(r"Piping\s*Plover\s*[-:]?\s*\(?\s*(\d+)\b", re.IGNORECASE),
]

def parse_pipl_count(text):
    if text is None:
        return 0
    s = str(text)
    for pat in _PIPL_COUNT_PATTERNS:
        m = pat.search(s)
        if m:
            try:
                return int(m.group(1))
            except ValueError:
                continue
    if re.search(r"\bPIPL\b|\bPiping\s*Plover\b", s, re.IGNORECASE):
        return 1
    return 0


_TRIVIAL_BAND_VALUES = {
    "", "0", "none", "n/a", "na", "-",
    "no", "no bands", "no banded birds", "none banded", "not banded",
}

def _is_trivial(cell):
    if cell is None:
        return True
    s = str(cell).strip().lower().rstrip(".")
    return s in _TRIVIAL_BAND_VALUES


def combine_band_cells(*cells):
    """Concatenate non-trivial band cells with ' | ' separator."""
    parts = []
    for c in cells:
        if c is None:
            continue
        s = str(c).strip()
        if not _is_trivial(s):
            parts.append(s)
    return " | ".join(parts) if parts else ""


def parse_band_entries(text):
    """Split a band text into individual bird entries (1) ... 2) ... etc.)"""
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return []
    text = str(text).strip()
    if not text or text.lower() in ("no banded birds", "none banded", "unbanded"):
        return []
    # Normalise inline numbering: "1) a 2) b" → "1) a\n2) b"
    text = re.sub(r'(?<!\n)(\s+)(\d+\)\s*)', r'\n\2', text)
    entries = []
    for line in text.split("\n"):
        line = line.strip()
        m = re.match(r'^\d+\)\s*', line)
        if m:
            entry = line[m.end():].strip()
            if entry:
                entries.append(entry)
    if not entries and text:
        entries = [text]
    return entries


# ══════════════════════════════════════════════════════════════════════════════
# Biologist correction resolution (same priority as 2018+ pipelines)
# ══════════════════════════════════════════════════════════════════════════════

def _normalize_correction_fmt(text: str) -> str:
    text = re.sub(r'(\d+\))\.\s*', r'\1 ', text)
    text = re.sub(r'\s+(\d+)\.\s+', lambda m: f'\n{m.group(1)}) ', text)
    return text.strip()


def _is_treat_as_unbanded(text: str) -> bool:
    if not text:
        return False
    t = text.lower()
    return "treat" in t and "unbanded" in t


def _resolve_band_text(notes: str, correction: str, interpretation: str):
    """Returns None=skip, ''=force unbanded, str=use this band text."""
    if correction:
        t = correction.lower().strip()
        if any(p in t for p in ("remove", "delete")):
            return None
    if _is_treat_as_unbanded(notes) or _is_treat_as_unbanded(correction):
        return ""
    if not correction:
        return interpretation or None
    t = correction.lower().strip()
    if ("unbanded" in t or "unreadable" in t) and "1)" not in t and "/" not in t:
        return ""
    if "1)" not in t and "/" not in t:
        if any(p in t for p in ("good", "confirm", "correct", "right")):
            return interpretation
    return _normalize_correction_fmt(correction)


def build_biologist_corrections():
    """
    Build {(normalized_route, date_str) → final_band_text} from the 2013
    sheet of the biologist review file.
    Returns {} if file/sheet missing.
    """
    if not review_path.exists():
        print("  [INFO] No biologist review file found — using raw band text")
        return {}
    wb = openpyxl.load_workbook(review_path, read_only=True)
    if config["biologist_review_sheet"] not in wb.sheetnames:
        print(f"  [INFO] No '{config['biologist_review_sheet']}' sheet — using raw band text")
        return {}
    ws = wb[config["biologist_review_sheet"]]

    corrections = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        route, point, date_str, pipl_at_pt, raw, interp, notes, correction = row[:8]
        if not route:
            continue
        text = _resolve_band_text(
            str(notes or ""), str(correction or ""), str(interp or "")
        )
        if text is None:
            continue
        key = (normalize_route(route), str(date_str or "").strip()[:10])
        corrections[key] = text

    wb.close()
    n_band     = sum(1 for v in corrections.values() if v)
    n_unbanded = sum(1 for v in corrections.values() if v == "")
    print(f"  Biologist corrections loaded: {n_band} band-text rows, "
          f"{n_unbanded} forced-unbanded rows")
    return corrections


# ══════════════════════════════════════════════════════════════════════════════
# Main extraction
# ══════════════════════════════════════════════════════════════════════════════

print(f"\n{'='*60}")
print(f"Step 0 — 2013 PIPL extraction")
print(f"  Source: {source_path}")
print(f"  Review: {review_path}")
print()

bio_corrections = build_biologist_corrections()

wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
ws = wb[config["sheet"]]

out_rows = []
stats = {
    "source_rows":    0,
    "rows_with_pipl": 0,
    "rows_banded":    0,
    "rows_unbanded":  0,
    "total_pipl":     0,
}

data_start = config["header_row"] + 1   # row after header
for row_tuple in ws.iter_rows(min_row=data_start, values_only=True):
    stats["source_rows"] += 1
    row = list(row_tuple) + [None] * 12   # pad to safe length
    date_val  = row[cols["date"]]
    route_raw = row[cols["route"]]
    species   = row[cols["species"]]
    lat       = parse_coord(row[cols["lat"]])
    lon       = parse_coord(row[cols["lon"]])
    observer  = row[cols["observer"]]
    email     = row[cols["email"]]
    flag      = row[cols["flag"]]
    combo     = row[cols["combo"]]
    comments  = row[cols["comments"]]

    if not route_raw:
        continue

    route = str(route_raw).strip()
    pipl  = parse_pipl_count(species)
    if pipl == 0:
        continue
    stats["rows_with_pipl"] += 1
    stats["total_pipl"] += pipl

    # Build raw band text from flag+combo cols
    raw_band = combine_band_cells(flag, combo)

    # Apply biologist correction ONLY if this source row originally had
    # non-trivial band content (otherwise the correction would wrongly
    # fan out to every same-(route,date) row that had no band data).
    if raw_band:
        key = (normalize_route(route), _format_date(date_val))
        if key in bio_corrections:
            raw_band = bio_corrections[key]

    # Parse to band entries
    band_list = parse_band_entries(raw_band) if raw_band else []
    n_banded  = len(band_list)
    remainder = pipl - n_banded

    if remainder < 0:
        print(f"  [WARN] {route} {_format_date(date_val)}: "
              f"{n_banded} banded > {pipl} PIPL — keeping all band rows")
        remainder = 0

    if pd.notna(date_val):
        try:
            date_val = pd.to_datetime(date_val)
        except Exception:
            pass

    base = {
        "SurveyDate":       date_val,
        "SurveyTime":       None,                # not captured in 2013
        "WeatherCondition": None,                # not captured in 2013
        "Route":            route,
        "Latitude":         lat,
        "Longitude":        lon,
        "GroupNumber":      None,                # no Point # column
        "Observer":         str(observer).strip() if observer else None,
        "ObserverEmail":    str(email).strip()    if email    else None,
        "FlagCode":         None,                # not split out for 2013
        "FlagColor":        None,                # not split out for 2013
        "Comments":         str(comments).strip() if comments else None,
    }

    # Banded rows (one per banded bird)
    for band_text in band_list:
        r = base.copy()
        r["TotalObserved"] = 1
        r["BandCombo"]     = band_text
        out_rows.append(r)
        stats["rows_banded"] += 1

    # Unbanded remainder
    if remainder > 0:
        r = base.copy()
        r["TotalObserved"] = remainder
        r["BandCombo"]     = None
        out_rows.append(r)
        stats["rows_unbanded"] += 1
    elif n_banded == 0:
        # No bands at all — single unbanded row with full PIPL count
        r = base.copy()
        r["TotalObserved"] = pipl
        r["BandCombo"]     = None
        out_rows.append(r)
        stats["rows_unbanded"] += 1

wb.close()

df_out = pd.DataFrame(out_rows)
df_out.to_excel(output_path, index=False)

print(f"\n{'─'*50}")
print(f"[DONE] Step 0 — 2013")
print(f"  Source rows processed     : {stats['source_rows']}")
print(f"  Rows with PIPL > 0        : {stats['rows_with_pipl']}")
print(f"  Output rows — banded      : {stats['rows_banded']}")
print(f"  Output rows — unbanded    : {stats['rows_unbanded']}")
print(f"  Total output rows         : {len(df_out)}")
print(f"  Total PIPL                : {stats['total_pipl']}")
print(f"  Extracted: {output_path}")
