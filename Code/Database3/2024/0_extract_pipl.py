"""
Step 0: Extract PIPL Data — 2024
──────────────────────────────────
Reads the 2024 Band Review file (wide format, one row per route) and melts
it into one row per bird per point, applying Option A band expansion.

Metadata (date, time, weather, observer, email) is NOT in the Band Review
sheet — it lives in the "All Species" sheet of Winter Birds 2024.xlsx
and is joined on the route name (col 0 = "Transect ").

Band data source priority per (route, point):
  1. Biologist correction  — Biologist_Band_Review_Completed_2024.xlsx col H
  2. Our interpretation    — same file col F (already structured 1)/2)/3))
  3. Raw Band Review text  — fallback if the review file has no entry for that point

_resolve_band_text logic (extends 2020/2021):
  • Confirmation phrases broadened to match any of "good" / "confirm" /
    "correct" / "right" (no "1)" / "/") → use our interpretation.
  • "treat as unbanded" directive honored in EITHER `Our Notes (read only)` or
    `Your Correction` column → forces unbanded.

2024-specific handling (mirrors 2020):
  ① Column names are short clean strings.
  ② Metadata joined from All Species sheet (skip the 'Totals' header row).
  ③ All 19 points present, all GPS columns present.
  ④ Observer email present in the form.
  ⑤ No survey-comments column.
  ⑥ GPS fallback: 2019, 2020, 2024–2024 clean files (skip 2024 itself).
  ⑦ Routes with parenthetical county suffix (e.g. "Anclote Key North (Pasco)")
    are preserved as-is — they match the All Species sheet exactly.
  ⑧ All Species column layout shifted in 2024 — Transect header expanded
    and a new "If you did not survey" column inserted. Handled by updating
    cols["route"] in the 2024 config; all other columns matched by name.

Output: db3_2024_extracted.xlsx  (in Output/2024/)
"""

import pandas as pd
import re
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))
from database3_config import config, get_output_folder, get_filename

script_dir    = os.path.dirname(os.path.abspath(__file__))
output_folder = get_output_folder(script_dir)
year          = config["year"]

review_dir   = os.path.normpath(os.path.join(script_dir, config["input_folder"]))
review_path  = os.path.join(review_dir, config["file"])
meta_dir     = os.path.normpath(os.path.join(script_dir, config["metadata_folder"]))
meta_path    = os.path.join(meta_dir, config["metadata_file"])
output_path  = os.path.join(output_folder, get_filename("extracted"))

points = config["points"]
cols   = config["columns"]


# ══════════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════════

def normalize_route(s):
    s = str(s).lower().strip()
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def strip_county(s):
    """Remove ', County=...' suffix from a route name (no-op for 2020 but harmless)."""
    return re.sub(r',\s*County=.*$', '', str(s), flags=re.IGNORECASE).strip()

def parse_coord(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        # Strip degree symbol and surrounding whitespace — some 2024 rows
        # (e.g. Biscayne National Park) store coords as '25.453172°'.
        s = str(val).strip().rstrip("°").strip()
        f = float(s)
        return f if f != 0.0 else None
    except (ValueError, TypeError):
        return None

def combine_weather(temp, wind, rain):
    parts = []
    for v in (temp, wind, rain):
        s = "" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v).strip()
        if s and s.lower() not in ("nan", "none", ""):
            parts.append(s)
    return ", ".join(parts) if parts else None

def parse_band_entries(text):
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return []
    text = str(text).strip()
    if not text or text.lower() == "no banded birds":
        return []
    # Normalise inline numbering: "1) bird1 2) bird2" → "1) bird1\n2) bird2"
    text = re.sub(r'(?<!\n)(\s+)(\d+\)\s*)', r'\n\2', text)
    entries = []
    for line in text.split("\n"):
        line = line.strip()
        m = re.match(r'^\d+\)\s*', line)
        if m:
            entry = line[m.end():].strip()
            if entry:
                entries.append(entry)
    if not entries and text:
        entries = [text]
    return entries


# ══════════════════════════════════════════════════════════════════════════════
# Biologist Corrections
# ══════════════════════════════════════════════════════════════════════════════

def _normalize_correction_fmt(text: str) -> str:
    """
    Fix common formatting issues biologists introduce when typing corrections:
      • '2). text'  →  '2) text'
      • ' 4. text'  →  '\\n4) text'
    """
    text = re.sub(r'(\d+\))\.\s*', r'\1 ', text)
    text = re.sub(r'\s+(\d+)\.\s+', lambda m: f'\n{m.group(1)}) ', text)
    return text.strip()


def _is_treat_as_unbanded(text: str) -> bool:
    """
    Detect 'treat ... as unbanded' / 'treat ... unbanded' directives.
    Checked against BOTH Our Notes and Your Correction per biologist guidance:
    when the band combo doesn't make sense, the birds are treated as unbanded.
    """
    if not text:
        return False
    t = text.lower()
    return "treat" in t and "unbanded" in t


def _resolve_band_text(notes: str, correction: str, interpretation: str):
    """
    Decide what band text to use given biologist notes, correction, and our interpretation.

    Returns:
      None  → skip this (route, point) entirely (e.g. 'remove this row')
      ''    → treat as no banded birds
      str   → the band text to parse (corrected, or confirmed interpretation)
    """
    # 1) "Remove / delete" → skip row
    if correction:
        t = correction.lower().strip()
        if any(p in t for p in ("remove", "delete")):
            return None

    # 2) "Treat as unbanded" directive — honored in either column
    if _is_treat_as_unbanded(notes) or _is_treat_as_unbanded(correction):
        return ""

    if not correction:
        return interpretation or None

    t = correction.lower().strip()

    # 3) "Unreadable / unbanded" instruction in correction
    if ("unbanded" in t or "unreadable" in t) and "1)" not in t and "/" not in t:
        return ""

    # 4) Confirmation phrases (broadened: good / confirm / correct / right)
    if "1)" not in t and "/" not in t:
        if any(p in t for p in ("good", "confirm", "correct", "right")):
            return interpretation

    # 5) Actual band data — normalise formatting then return
    return _normalize_correction_fmt(correction)


def build_biologist_corrections():
    """
    Read the completed biologist review file (2020 sheet) and build:
        (normalized_route, point_N) → final_band_text
    Returns {} gracefully if file or sheet is missing.
    """
    review_path_bio = (
        Path(script_dir).parents[2]
        / "Databases" / "Database3BiologistReview"
        / "Biologist_Band_Review_Completed_2024.xlsx"
    )
    if not review_path_bio.exists():
        print("  [INFO] No biologist review file found — using raw Band Review data")
        return {}

    import openpyxl
    wb = openpyxl.load_workbook(review_path_bio)
    if year not in wb.sheetnames:
        print(f"  [INFO] No '{year}' sheet in biologist review file — using raw Band Review data")
        return {}
    ws  = wb[year]
    corrections = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        route      = str(row[0] or "").strip()
        point      = row[1]
        interp     = str(row[5] or "").strip()   # col F — our interpretation
        notes      = str(row[6] or "").strip()   # col G — our notes (read only)
        correction = str(row[7] or "").strip()   # col H — biologist correction

        if not route or not point:
            continue

        text = _resolve_band_text(notes, correction, interp)
        if text is None:
            continue   # skip row (delete instruction)

        key = (normalize_route(strip_county(route)), int(point))
        corrections[key] = text

    n_band     = sum(1 for v in corrections.values() if v)
    n_unbanded = sum(1 for v in corrections.values() if v == "")
    print(f"  Biologist corrections loaded: {n_band} band entries, "
          f"{n_unbanded} points overridden to unbanded")
    return corrections


# ══════════════════════════════════════════════════════════════════════════════
# Metadata join — read All Species sheet of the raw 2020 file
# ══════════════════════════════════════════════════════════════════════════════

def build_metadata_lookup():
    """
    Build a lookup keyed by normalized route name:
        normalized_route → dict of {date, time, temp, wind, rain, observer, email}
    Skips the TOTALS row at the top of the All Species sheet.

    Also adds entries for any aliases in config["route_aliases"] so that
    Band Review names that differ from All Species names still join.
    """
    df_meta = pd.read_excel(meta_path, sheet_name=config["metadata_sheet"], header=0)
    print(f"  Loaded {len(df_meta)} rows from All Species sheet")

    route_col_name = cols["route"]

    lookup = {}
    skipped_totals = 0
    for _, r in df_meta.iterrows():
        raw_route = r.get(route_col_name)
        if raw_route is None or (isinstance(raw_route, float) and pd.isna(raw_route)):
            continue
        if str(raw_route).strip().lower() in ("totals", "total"):
            skipped_totals += 1
            continue

        key = normalize_route(strip_county(str(raw_route)))
        if key in lookup:
            continue  # first wins on duplicate route names

        lookup[key] = {
            "date":     r.get(cols["date"]),
            "time":     r.get(cols["time"]),
            "temp":     r.get(cols["temp"]),
            "wind":     r.get(cols["wind"]),
            "rain":     r.get(cols["rain"]),
            "observer": r.get(cols["observer"]),
            "email":    r.get(cols["email"]),
        }

    if skipped_totals:
        print(f"  Skipped {skipped_totals} TOTALS row(s) in All Species")

    # ── Apply route aliases: BR-name → AS-name ────────────────────────────────
    aliases = config.get("route_aliases", {})
    alias_hits = 0
    for br_name, as_name in aliases.items():
        as_key = normalize_route(strip_county(as_name))
        br_key = normalize_route(strip_county(br_name))
        if br_key in lookup:
            continue  # already joins directly
        if as_key in lookup:
            lookup[br_key] = lookup[as_key]
            alias_hits += 1
            print(f"  [ALIAS] {br_name!r}")
            print(f"          → joined via {as_name!r}")
    if aliases:
        print(f"  Route aliases applied: {alias_hits}/{len(aliases)}")

    print(f"  Metadata lookup: {len(lookup)} routes")
    return lookup


# ══════════════════════════════════════════════════════════════════════════════
# GPS fallback from other years (skip current year)
# ══════════════════════════════════════════════════════════════════════════════

def build_gps_fallback():
    """
    Scan clean Focal Observations sheets for years != current year and build:
        (normalized_route, point_N) → (lat, lon, source_year)
    """
    lookup = {}
    root   = Path(script_dir).parents[2] / "Databases" / "Database3Clean"

    for yr in range(2019, 2025):
        if str(yr) == year:
            continue
        fpath = root / f"Winter Birds {yr} Clean.xlsx"
        if not fpath.exists():
            continue
        try:
            wb_tmp = pd.ExcelFile(fpath)
        except Exception:
            continue
        if "Focal Observations" not in wb_tmp.sheet_names:
            continue
        df_tmp = wb_tmp.parse("Focal Observations", header=0)
        route_col_name = df_tmp.columns[0]

        for n in range(1, 20):
            lat_col = next((c for c in df_tmp.columns
                            if re.search(rf'Lat.*\bpoint\s+{n}\b', str(c), re.IGNORECASE)
                            or re.search(rf'\bpoint\s+{n}\b.*Lat', str(c), re.IGNORECASE)), None)
            lon_col = next((c for c in df_tmp.columns
                            if re.search(rf'Long.*\bpoint\s+{n}\b', str(c), re.IGNORECASE)
                            or re.search(rf'\bpoint\s+{n}\b.*Long', str(c), re.IGNORECASE)), None)
            if lat_col is None or lon_col is None:
                continue
            for _, r in df_tmp.iterrows():
                lat = parse_coord(r.get(lat_col))
                lon = parse_coord(r.get(lon_col))
                route_val = r.get(route_col_name)
                if route_val and lat and lon:
                    key = (normalize_route(strip_county(str(route_val))), int(n))
                    if key not in lookup:
                        lookup[key] = (lat, lon, yr)

    print(f"  GPS fallback lookup: {len(lookup)} (route, point) pairs from other years")
    return lookup


# ══════════════════════════════════════════════════════════════════════════════
# Map column indices from Band Review headers
# ══════════════════════════════════════════════════════════════════════════════

def build_point_col_map(headers):
    """
    Returns dict: point_N → {lat, long, pipl, band} column indices.
    Matches the short 2020 header format: 'Latitude (point N)' etc.
    """
    point_map = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        h_str = str(h)

        m = re.search(r'Latitude.*point\s+(\d+)', h_str, re.IGNORECASE)
        if m:
            n = int(m.group(1))
            point_map.setdefault(n, {})["lat"] = i
            continue

        m = re.search(r'Longitude.*point\s+(\d+)', h_str, re.IGNORECASE)
        if m:
            n = int(m.group(1))
            point_map.setdefault(n, {})["long"] = i
            continue

        m = re.search(r'Number.*Piping.*point\s+(\d+)', h_str, re.IGNORECASE)
        if m:
            n = int(m.group(1))
            point_map.setdefault(n, {})["pipl"] = i
            continue

        m = re.search(r'PIPL.*point\s+(\d+)', h_str, re.IGNORECASE)
        if m:
            n = int(m.group(1))
            point_map.setdefault(n, {})["band"] = i

    return point_map


# ══════════════════════════════════════════════════════════════════════════════
# Main extraction
# ══════════════════════════════════════════════════════════════════════════════

print(f"\n{'='*60}")
print(f"Step 0 — 2024 PIPL extraction")
print(f"  Review:   {review_path}")
print(f"  Metadata: {meta_path}")
print()

df_raw = pd.read_excel(review_path, sheet_name=config["sheet"], header=config["header_row"])
print(f"  Loaded {len(df_raw)} route rows from Band Review")

# ── Load biologist corrections ─────────────────────────────────────────────────
bio_corrections = build_biologist_corrections()

# ── Build metadata lookup from All Species ─────────────────────────────────────
metadata = build_metadata_lookup()

# ── Build GPS fallback from other years ────────────────────────────────────────
gps_fallback = build_gps_fallback()

# ── Map point columns ──────────────────────────────────────────────────────────
headers_list = list(df_raw.columns)
point_cols   = build_point_col_map(headers_list)

route_col_idx = 0   # col 0 is the Transect column in Band Review

# Pre-pipeline removed rows (none expected for 2020, but keep the slot for parity)
pre_removed_rows = []

# ── Melt wide → long with Option A expansion ──────────────────────────────────
out_rows = []

stats = {
    "routes":                 len(df_raw),
    "points_skipped_no_pipl": 0,
    "points_dropped_no_gps":  0,
    "points_gps_fallback":    0,
    "rows_banded":            0,
    "rows_unbanded":          0,
    "total_pipl":             0,
    "routes_no_metadata":     0,
}

missing_meta_routes = set()

for _, route_row in df_raw.iterrows():

    route_raw = str(route_row.iloc[route_col_idx] or "").strip()
    if not route_raw:
        continue
    route     = strip_county(route_raw)
    route_key = normalize_route(route)

    meta = metadata.get(route_key, {})
    if not meta:
        stats["routes_no_metadata"] += 1
        missing_meta_routes.add(route)

    date_val = meta.get("date")
    time_val = meta.get("time")
    weather  = combine_weather(meta.get("temp"), meta.get("wind"), meta.get("rain"))
    observer = meta.get("observer")
    email    = meta.get("email")

    if pd.notna(date_val):
        try:
            date_val = pd.to_datetime(date_val)
        except Exception:
            pass

    for n in points:
        pc = point_cols.get(n, {})

        # PIPL count
        pipl_raw = route_row.iloc[pc["pipl"]] if "pipl" in pc else None
        try:
            pipl_count = int(float(pipl_raw)) if pipl_raw is not None and not (
                isinstance(pipl_raw, float) and pd.isna(pipl_raw)) else 0
        except (ValueError, TypeError):
            pipl_count = 0

        if pipl_count == 0:
            stats["points_skipped_no_pipl"] += 1
            continue

        # Latitude / Longitude
        lat = parse_coord(route_row.iloc[pc["lat"]])  if "lat"  in pc else None
        lon = parse_coord(route_row.iloc[pc["long"]]) if "long" in pc else None

        gps_note = None

        # GPS fallback from other years
        if lat is None or lon is None:
            fb_key = (route_key, n)
            if fb_key in gps_fallback:
                fb_lat, fb_lon, fb_yr = gps_fallback[fb_key]
                if lat is None:
                    lat = fb_lat
                if lon is None:
                    lon = fb_lon
                stats["points_gps_fallback"] += 1
                gps_note = f"GPS borrowed from {fb_yr} — point {n} had no coordinates in 2024"

        if lat is None or lon is None:
            stats["points_dropped_no_gps"] += 1
            out_rows.append({
                "SurveyDate":      date_val,
                "Route":           route,
                "GroupNumber":     n,
                "TotalObserved":   pipl_count,
                "BandCombo":       None,
                "_removal_reason": f"No GPS for point {n} — not found in raw file or any other year",
            })
            print(f"  [DROP] {route} pt {n}: PIPL={pipl_count}, no GPS in any source")
            continue

        # Band entries — biologist correction first, else raw Band Review text
        band_raw = route_row.iloc[pc["band"]] if "band" in pc else None
        bio_key  = (route_key, n)
        if bio_key in bio_corrections:
            band_raw = bio_corrections[bio_key]   # "" = unbanded; string = corrected combo
        band_list = parse_band_entries(band_raw)
        n_banded  = len(band_list)
        remainder = pipl_count - n_banded

        if remainder < 0:
            print(f"  [WARNING] {route} pt {n}: {n_banded} banded entries > PIPL count "
                  f"({pipl_count}). Keeping all band rows — flag for biologist review.")
            remainder = 0

        base = {
            "SurveyDate":       date_val,
            "SurveyTime":       str(time_val).strip() if pd.notna(time_val) else None,
            "WeatherCondition": weather,
            "Route":            route,
            "Latitude":         lat,
            "Longitude":        lon,
            "GroupNumber":      n,
            "Observer":         str(observer).strip() if pd.notna(observer) else None,
            "ObserverEmail":    str(email).strip()    if pd.notna(email)    else None,
            "FlagCode":         None,
            "FlagColor":        None,
            "Comments":         None,  # no survey-comments column in 2020 form
        }

        # Banded rows
        for band_text in band_list:
            row = base.copy()
            row["TotalObserved"] = 1
            row["BandCombo"]     = band_text
            out_rows.append(row)
            stats["rows_banded"] += 1

        # Unbanded remainder
        if remainder > 0:
            row = base.copy()
            row["TotalObserved"] = remainder
            row["BandCombo"]     = None
            out_rows.append(row)
            stats["rows_unbanded"] += 1
        elif n_banded == 0:
            row = base.copy()
            row["TotalObserved"] = pipl_count
            row["BandCombo"]     = None
            out_rows.append(row)
            stats["rows_unbanded"] += 1

        stats["total_pipl"] += pipl_count

# ── Save ──────────────────────────────────────────────────────────────────────
all_rows = out_rows + pre_removed_rows
df_out   = pd.DataFrame(all_rows)
df_out.to_excel(output_path, index=False)

print(f"\n{'─'*50}")
print(f"[DONE] Step 0 — 2024")
print(f"  Routes processed:             {stats['routes']}")
print(f"  Points skipped (0 PIPL):      {stats['points_skipped_no_pipl']}")
print(f"  Points GPS from fallback:     {stats['points_gps_fallback']}")
print(f"  Points dropped (no GPS):      {stats['points_dropped_no_gps']}")
print(f"  Routes with no metadata:      {stats['routes_no_metadata']}")
if missing_meta_routes:
    print(f"  [WARN] Routes missing All Species metadata join:")
    for r in sorted(missing_meta_routes):
        print(f"         - {r}")
print(f"  Output rows — banded:         {stats['rows_banded']}")
print(f"  Output rows — unbanded:       {stats['rows_unbanded']}")
print(f"  Total output rows:            {len(df_out)}")
print(f"  Total PIPL:                   {stats['total_pipl']}")
print(f"  Extracted: {output_path}")
