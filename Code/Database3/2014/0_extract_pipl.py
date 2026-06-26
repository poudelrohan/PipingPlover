"""
Step 0: Extract PIPL Data — 2014
──────────────────────────────────
Reads Winter Birds '14 Clean.xlsx ('Indiv Flock GPS & bands' sheet) and
emits one row per individual bird (Option A expansion), applying biologist
corrections from Biologist_Band_Review_2013-2017.xlsx sheet "2014".

2014-specific handling:
  - GroupNumber sourced from col 3 ("Group or Point #").
  - PIPL count extracted from col 6 free-text species column
    (e.g. "PIPL (5 TOTAL, 3 banded);  DUNL (45)").
  - Band info is a single column (col 7). Goes into BandCombo
    (fixes prior bug where band info ended up in FlagCode column).
  - FlagCode and FlagColor stay null in 2014 — no structured fields exist.
  - Biologist corrections keyed by (route, date, point).

Output: db3_2014_extracted.xlsx in Output/2014/
"""

import pandas as pd
import re
import os
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
from database3_config import config, get_output_folder, get_filename

script_dir    = os.path.dirname(os.path.abspath(__file__))
output_folder = get_output_folder(script_dir)
year          = config["year"]

source_path = os.path.normpath(
    os.path.join(script_dir, config["input_folder"], config["file"])
)
review_path = (
    Path(script_dir).parents[2] / "Databases" / "Database3BiologistReview"
    / config["biologist_review_file"]
)
output_path = os.path.join(output_folder, get_filename("extracted"))

cols = config["columns"]


# ══════════════════════════════════════════════════════════════════════════════
# Helpers (same shape as 2013 / 2018+ pipelines)
# ══════════════════════════════════════════════════════════════════════════════

def normalize_route(s):
    s = str(s or "").lower().strip()
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def _format_date(dv):
    if dv is None:
        return ""
    if hasattr(dv, "strftime"):
        try:
            return dv.strftime("%Y-%m-%d")
        except Exception:
            pass
    return str(dv)[:10]


def parse_coord(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        s = str(val).strip().rstrip("°").strip()
        f = float(s)
        return f if f != 0.0 else None
    except (ValueError, TypeError):
        return None


_PIPL_COUNT_PATTERNS = [
    re.compile(r"PIPL\s*[-:]?\s*\(?\s*(\d+)\b",            re.IGNORECASE),
    re.compile(r"(\d+)\s*PIPL\b",                          re.IGNORECASE),
    re.compile(r"PIPL\s*\(\s*(\d+)\b",                     re.IGNORECASE),
    re.compile(r"Piping\s*Plover\s*[-:]?\s*\(?\s*(\d+)\b", re.IGNORECASE),
]

def parse_pipl_count(text):
    if text is None:
        return 0
    s = str(text)
    for pat in _PIPL_COUNT_PATTERNS:
        m = pat.search(s)
        if m:
            try:
                return int(m.group(1))
            except ValueError:
                continue
    if re.search(r"\bPIPL\b|\bPiping\s*Plover\b", s, re.IGNORECASE):
        return 1
    return 0


_TRIVIAL_BAND_VALUES = {
    "", "0", "none", "n/a", "na", "-",
    "no", "no bands", "no banded birds", "none banded", "not banded",
    "no bands observed", "bands unreadable",
}

def _is_trivial(cell):
    if cell is None:
        return True
    s = str(cell).strip().lower().rstrip(".")
    return s in _TRIVIAL_BAND_VALUES


def parse_band_entries(text):
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return []
    text = str(text).strip()
    if not text or text.lower() in ("no banded birds", "none banded", "unbanded"):
        return []
    text = re.sub(r'(?<!\n)(\s+)(\d+\)\s*)', r'\n\2', text)
    entries = []
    for line in text.split("\n"):
        line = line.strip()
        m = re.match(r'^\d+\)\s*', line)
        if m:
            entry = line[m.end():].strip()
            if entry:
                entries.append(entry)
    if not entries and text:
        entries = [text]
    return entries


# ══════════════════════════════════════════════════════════════════════════════
# Biologist correction resolution
# ══════════════════════════════════════════════════════════════════════════════

def _normalize_correction_fmt(text: str) -> str:
    text = re.sub(r'(\d+\))\.\s*', r'\1 ', text)
    text = re.sub(r'\s+(\d+)\.\s+', lambda m: f'\n{m.group(1)}) ', text)
    return text.strip()


def _is_treat_as_unbanded(text: str) -> bool:
    if not text:
        return False
    t = text.lower()
    return "treat" in t and "unbanded" in t


def _resolve_band_text(notes: str, correction: str, interpretation: str):
    if correction:
        t = correction.lower().strip()
        if any(p in t for p in ("remove", "delete")):
            return None
    if _is_treat_as_unbanded(notes) or _is_treat_as_unbanded(correction):
        return ""
    if not correction:
        return interpretation or None
    t = correction.lower().strip()
    if ("unbanded" in t or "unreadable" in t) and "1)" not in t and "/" not in t:
        return ""
    if "1)" not in t and "/" not in t:
        if any(p in t for p in ("good", "confirm", "correct", "right")):
            return interpretation
    return _normalize_correction_fmt(correction)


def build_biologist_corrections():
    """
    Build {(normalized_route, date_str, point_str) → final_band_text}
    from the 2014 sheet of the biologist review file.
    """
    if not review_path.exists():
        print("  [INFO] No biologist review file found — using raw band text")
        return {}
    wb = openpyxl.load_workbook(review_path, read_only=True)
    sheet_name = config["biologist_review_sheet"]
    if sheet_name not in wb.sheetnames:
        print(f"  [INFO] No '{sheet_name}' sheet — using raw band text")
        return {}
    ws = wb[sheet_name]

    corrections = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        route, point, date_str, pipl_at_pt, raw, interp, notes, correction = row[:8]
        if not route:
            continue
        text = _resolve_band_text(
            str(notes or ""), str(correction or ""), str(interp or "")
        )
        if text is None:
            continue
        key = (
            normalize_route(route),
            str(date_str or "").strip()[:10],
            str(point).strip() if point is not None else "",
        )
        corrections[key] = text

    wb.close()
    n_band     = sum(1 for v in corrections.values() if v)
    n_unbanded = sum(1 for v in corrections.values() if v == "")
    print(f"  Biologist corrections loaded: {n_band} band-text rows, "
          f"{n_unbanded} forced-unbanded rows")
    return corrections


# ══════════════════════════════════════════════════════════════════════════════
# Main extraction
# ══════════════════════════════════════════════════════════════════════════════

print(f"\n{'='*60}")
print(f"Step 0 — 2014 PIPL extraction")
print(f"  Source: {source_path}")
print(f"  Review: {review_path}")
print()

bio_corrections = build_biologist_corrections()

wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
ws = wb[config["sheet"]]

out_rows = []
stats = {
    "source_rows":    0,
    "rows_with_pipl": 0,
    "rows_banded":    0,
    "rows_unbanded":  0,
    "total_pipl":     0,
}

data_start = config["header_row"] + 1
for row_tuple in ws.iter_rows(min_row=data_start, values_only=True):
    stats["source_rows"] += 1
    row = list(row_tuple) + [None] * 10
    date_val  = row[cols["date"]]
    route_raw = row[cols["route"]]
    observer  = row[cols["observer"]]
    point     = row[cols["point"]]
    lat       = parse_coord(row[cols["lat"]])
    lon       = parse_coord(row[cols["lon"]])
    species   = row[cols["species"]]
    bands     = row[cols["bands"]]
    notes_col = row[cols["notes"]]

    if not route_raw:
        continue
    route = str(route_raw).strip()
    pipl  = parse_pipl_count(species)
    if pipl == 0:
        continue
    stats["rows_with_pipl"] += 1
    stats["total_pipl"] += pipl

    # Trivial band → empty raw
    raw_band = "" if _is_trivial(bands) else str(bands).strip()

    # Apply biologist correction ONLY if this source row originally had
    # non-trivial band content (avoids spurious overrides on no-band rows
    # that happen to share the same route/date/point key).
    if raw_band:
        key = (
            normalize_route(route),
            _format_date(date_val),
            str(point).strip() if point is not None else "",
        )
        if key in bio_corrections:
            raw_band = bio_corrections[key]

    band_list = parse_band_entries(raw_band) if raw_band else []
    n_banded  = len(band_list)
    remainder = pipl - n_banded

    if remainder < 0:
        print(f"  [WARN] {route} pt {point} {_format_date(date_val)}: "
              f"{n_banded} banded > {pipl} PIPL — keeping all band rows")
        remainder = 0

    if pd.notna(date_val):
        try:
            date_val = pd.to_datetime(date_val)
        except Exception:
            pass

    base = {
        "SurveyDate":       date_val,
        "SurveyTime":       None,
        "WeatherCondition": None,
        "Route":            route,
        "Latitude":         lat,
        "Longitude":        lon,
        "GroupNumber":      point if (point is not None and str(point).strip()) else None,
        "Observer":         str(observer).strip() if observer else None,
        "ObserverEmail":    None,                                 # 2014 has no email column
        "FlagCode":         None,                                 # not used in 2014
        "FlagColor":        None,                                 # not used in 2014
        "Comments":         str(notes_col).strip() if notes_col and not (isinstance(notes_col, float) and pd.isna(notes_col)) else None,
    }

    for band_text in band_list:
        r = base.copy()
        r["TotalObserved"] = 1
        r["BandCombo"]     = band_text
        out_rows.append(r)
        stats["rows_banded"] += 1

    if remainder > 0:
        r = base.copy()
        r["TotalObserved"] = remainder
        r["BandCombo"]     = None
        out_rows.append(r)
        stats["rows_unbanded"] += 1
    elif n_banded == 0:
        r = base.copy()
        r["TotalObserved"] = pipl
        r["BandCombo"]     = None
        out_rows.append(r)
        stats["rows_unbanded"] += 1

wb.close()

df_out = pd.DataFrame(out_rows)
df_out.to_excel(output_path, index=False)

print(f"\n{'─'*50}")
print(f"[DONE] Step 0 — 2014")
print(f"  Source rows processed     : {stats['source_rows']}")
print(f"  Rows with PIPL > 0        : {stats['rows_with_pipl']}")
print(f"  Output rows — banded      : {stats['rows_banded']}")
print(f"  Output rows — unbanded    : {stats['rows_unbanded']}")
print(f"  Total output rows         : {len(df_out)}")
print(f"  Total PIPL                : {stats['total_pipl']}")
print(f"  Extracted: {output_path}")
