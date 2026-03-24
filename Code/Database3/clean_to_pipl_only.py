"""
Database 3 — Strip yearly Winter Bird files to PIPL-only data.
═══════════════════════════════════════════════════════════════════
Copies the original Excel files and modifies them IN-PLACE using openpyxl
so that all original formatting, colors, column widths, and styles are preserved.

Only rows/columns are deleted — everything else stays exactly as the original.

  Datasheet 1 (species+GPS):
    Delete rows where the species/focal column does NOT mention "PIPL".

  Datasheet 2 (survey count columns):
    Delete all species count columns EXCEPT Piping Plover.
    Delete rows where Piping Plover is blank or 0.

  Datasheet 3 (individual species — 2015/2016 only):
    Delete rows where Species column does NOT contain "PIPL".

Output: Databases/Database3Clean/<filename> Clean.xlsx
"""

import shutil
import os
import re
from openpyxl import load_workbook

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR   = os.path.normpath(os.path.join(SCRIPT_DIR, "../../Databases/Database3"))
OUTPUT_DIR  = os.path.normpath(os.path.join(SCRIPT_DIR, "../../Databases/Database3Clean"))
os.makedirs(OUTPUT_DIR, exist_ok=True)

PIPL_PATTERN = re.compile(r"PIPL|[Pp]iping\s*[Pp]lover", re.IGNORECASE)


# ── Year-specific sheet configs ───────────────────────────────────────────────
YEAR_CONFIGS = {
    "Winter Birds '13.xlsx": {
        "datasheet1": {
            "sheet": "Species GPS",
            "header_row": 2,          # 1-indexed in openpyxl (row 1 = annotations, row 2 = headers)
            "data_start_row": 3,      # first data row
            "species_col_name": "Focal species",  # partial match on header text
        },
        "datasheet2": {
            "sheet": "counts",
            "header_row": 2,
            "data_start_row": 3,
            "pipl_col_name": "Piping Plover",
            "metadata_end_col_name": "Starting Time",
        },
        "datasheet3": None,
    },
    "Winter Birds '14.xlsx": {
        "datasheet1": {
            "sheet": "Indiv Flock GPS & bands",
            "header_row": 2,
            "data_start_row": 3,
            "species_col_name": "Species and number",  # partial match
        },
        "datasheet2": {
            "sheet": "Total Survey Counts",
            "header_row": 1,          # 2014 has headers at row 1
            "data_start_row": 2,
            "pipl_col_name": "Piping Plover",
            "metadata_end_col_name": "Starting Time",
        },
        "datasheet3": None,
    },
    "Winter Birds '15.xlsx": {
        "datasheet1": {
            "sheet": "DATA SHEET 1",
            "header_row": 2,
            "data_start_row": 3,
            "species_col_name": "Species and number",
        },
        "datasheet2": {
            "sheet": "DATA SHEET 2",
            "header_row": 2,
            "data_start_row": 3,
            "pipl_col_name": "Piping Plover",
            "metadata_end_col_name": "Weather Condition",
        },
        "datasheet3": {
            "sheet": "DATA SHEET 3",
            "header_row": 2,
            "data_start_row": 3,
            "species_col_name": "Species",
        },
    },
    "Winter Birds '16.xlsx": {
        "datasheet1": {
            "sheet": "DATA SHEET 1",
            "header_row": 2,
            "data_start_row": 3,
            "species_col_name": "Species and number",
        },
        "datasheet2": {
            "sheet": "DATA SHEET 2",
            "header_row": 2,
            "data_start_row": 3,
            "pipl_col_name": "Piping Plover",
            "metadata_end_col_name": "Weather Condition",
        },
        "datasheet3": {
            "sheet": "DATA SHEET 3",
            "header_row": 2,
            "data_start_row": 3,
            "species_col_name": "Species",
        },
    },
}


def find_col_by_name(ws, header_row, partial_name):
    """Find column index (1-indexed) by partial header text match."""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and partial_name.lower() in str(val).lower():
            return col
    return None


def is_row_empty(ws, row):
    """Check if all cells in a row are None/empty."""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None and str(val).strip() != "":
            return False
    return True


def cell_mentions_pipl(val):
    """Check if a cell value mentions PIPL."""
    if val is None:
        return False
    return bool(PIPL_PATTERN.search(str(val)))


def clean_datasheet1(ws, cfg):
    """Delete rows where species column doesn't mention PIPL."""
    header_row = cfg["header_row"]
    data_start = cfg["data_start_row"]
    species_col = find_col_by_name(ws, header_row, cfg["species_col_name"])

    if species_col is None:
        print(f"    [WARNING] Could not find species column '{cfg['species_col_name']}'")
        return 0

    # Collect rows to delete (bottom-up to preserve indices)
    rows_to_delete = []
    kept = 0
    for row in range(ws.max_row, data_start - 1, -1):
        # Skip completely empty rows
        if is_row_empty(ws, row):
            rows_to_delete.append(row)
            continue

        val = ws.cell(row=row, column=species_col).value
        if not cell_mentions_pipl(val):
            rows_to_delete.append(row)
        else:
            kept += 1

    for row in rows_to_delete:
        ws.delete_rows(row)

    return kept


def clean_datasheet2(ws, cfg):
    """
    1) Delete all species count columns except Piping Plover
    2) Delete rows where Piping Plover is blank or 0
    """
    header_row = cfg["header_row"]
    data_start = cfg["data_start_row"]

    # Find the Piping Plover column and metadata boundary
    pipl_col = find_col_by_name(ws, header_row, cfg["pipl_col_name"])
    meta_end_col = find_col_by_name(ws, header_row, cfg["metadata_end_col_name"])

    if pipl_col is None:
        print(f"    [WARNING] Could not find PIPL column '{cfg['pipl_col_name']}'")
        return 0, 0

    if meta_end_col is None:
        # Fallback: assume first 11 columns are metadata
        meta_end_col = 11
        print(f"    [NOTE] Metadata end col not found, using column 11 as boundary")

    # Also find Comments and GRAND TOTAL columns to keep
    comments_col = find_col_by_name(ws, header_row, "Comments")
    grand_total_col = find_col_by_name(ws, header_row, "GRAND TOTAL")

    # Identify species columns to delete (between metadata and end, excluding PIPL)
    # Work right-to-left to preserve column indices
    cols_to_delete = []
    for col in range(ws.max_column, meta_end_col, -1):
        if col == pipl_col:
            continue
        if col == comments_col:
            continue
        if col == grand_total_col:
            continue
        # This is a species column (not PIPL, not Comments, not GRAND TOTAL)
        cols_to_delete.append(col)

    # Delete species columns (right to left)
    for col in cols_to_delete:
        ws.delete_cols(col)

    # Re-find PIPL column after deletions (it may have shifted)
    pipl_col = find_col_by_name(ws, header_row, cfg["pipl_col_name"])

    # Now delete rows where Piping Plover is blank or 0
    rows_to_delete = []
    kept = 0
    for row in range(ws.max_row, data_start - 1, -1):
        if is_row_empty(ws, row):
            rows_to_delete.append(row)
            continue

        # Check if this is a "Leave blank" instruction row
        first_val = ws.cell(row=row, column=1).value
        if first_val and "leave blank" in str(first_val).lower():
            rows_to_delete.append(row)
            continue

        val = ws.cell(row=row, column=pipl_col).value
        try:
            num_val = float(val) if val is not None else 0
        except (ValueError, TypeError):
            num_val = 0

        if num_val <= 0:
            rows_to_delete.append(row)
        else:
            kept += 1

    for row in rows_to_delete:
        ws.delete_rows(row)

    remaining_cols = ws.max_column
    return kept, remaining_cols


def clean_datasheet3(ws, cfg):
    """Delete rows where Species column doesn't contain PIPL."""
    header_row = cfg["header_row"]
    data_start = cfg["data_start_row"]
    species_col = find_col_by_name(ws, header_row, cfg["species_col_name"])

    if species_col is None:
        print(f"    [WARNING] Could not find species column '{cfg['species_col_name']}'")
        return 0

    rows_to_delete = []
    kept = 0
    for row in range(ws.max_row, data_start - 1, -1):
        if is_row_empty(ws, row):
            rows_to_delete.append(row)
            continue

        val = ws.cell(row=row, column=species_col).value
        if not cell_mentions_pipl(val):
            rows_to_delete.append(row)
        else:
            kept += 1

    for row in rows_to_delete:
        ws.delete_rows(row)

    return kept


# ── Main loop ─────────────────────────────────────────────────────────────────
for filename, sheets_cfg in YEAR_CONFIGS.items():
    filepath = os.path.join(INPUT_DIR, filename)

    if not os.path.exists(filepath):
        print(f"[SKIP] File not found: {filename}")
        continue

    # Build output filename: "Winter Birds '13.xlsx" → "Winter Birds '13 Clean.xlsx"
    base, ext = os.path.splitext(filename)
    out_filename = f"{base} Clean{ext}"
    output_path = os.path.join(OUTPUT_DIR, out_filename)

    # Step 1: Copy original file
    shutil.copy2(filepath, output_path)

    print(f"\n{'=' * 60}")
    print(f"Processing: {filename} → {out_filename}")
    print(f"{'=' * 60}")

    # Step 2: Open copy with openpyxl and modify in-place
    wb = load_workbook(output_path)

    # ── Datasheet 1 ──────────────────────────────────────────────
    ds1_cfg = sheets_cfg["datasheet1"]
    ws1 = wb[ds1_cfg["sheet"]]
    print(f"\n  Datasheet 1: '{ds1_cfg['sheet']}'")
    kept1 = clean_datasheet1(ws1, ds1_cfg)
    print(f"    → {kept1} PIPL rows kept")

    # ── Datasheet 2 ──────────────────────────────────────────────
    ds2_cfg = sheets_cfg["datasheet2"]
    ws2 = wb[ds2_cfg["sheet"]]
    print(f"\n  Datasheet 2: '{ds2_cfg['sheet']}'")
    kept2, cols2 = clean_datasheet2(ws2, ds2_cfg)
    print(f"    → {kept2} PIPL rows kept, {cols2} columns remaining")

    # ── Datasheet 3 (2015/2016 only) ─────────────────────────────
    ds3_cfg = sheets_cfg["datasheet3"]
    if ds3_cfg is not None:
        ws3 = wb[ds3_cfg["sheet"]]
        print(f"\n  Datasheet 3: '{ds3_cfg['sheet']}'")
        kept3 = clean_datasheet3(ws3, ds3_cfg)
        print(f"    → {kept3} PIPL rows kept")
    else:
        print(f"\n  Datasheet 3: N/A for this year")

    # ── Delete Introduction/README sheet ─────────────────────────
    intro_names = ["Introduction", "READ ME FIRST"]
    for name in intro_names:
        if name in wb.sheetnames:
            del wb[name]
            print(f"\n  Removed '{name}' sheet")

    # Step 3: Save
    wb.save(output_path)
    wb.close()
    print(f"\n  ✓ Saved: {out_filename}")
