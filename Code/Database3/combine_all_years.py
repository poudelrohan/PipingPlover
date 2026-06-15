"""
Combine all 12 years of Database 3 (Winter Bird Survey) into one folder for
biologist review.

Input : Databases/Database3/Output/<year>/db3_<year>_FINAL.xlsx (2013–2024)

Output:
  Databases/Database3/Output/AllYears/
    db3_2013_FINAL.xlsx        ← copies of each year's per-year FINAL file
    db3_2014_FINAL.xlsx
    ...
    db3_2024_FINAL.xlsx
    db3_ALL_YEARS_FINAL.xlsx   ← combined workbook (see sheets below)

Sheets in db3_ALL_YEARS_FINAL.xlsx:
  1. All_Years_Combined  — every year's clean rows stacked, 1,093 rows,
                           with new ID columns at the front.
  2. Summary             — one row per year of stats (routes, rows, PIPL,
                           banded, unbanded, % banded) plus a TOTAL row.
  3. 2013, 2014, ..., 2024 — one sheet per year showing just that year's
                           rows from the combined view (easy click-through).

Columns added by this script to All_Years_Combined and the per-year sheets:
  unique_id   : new global sequential int (1..N)
  database    : 3   (room for DB1/2/4 later)
  year_id     : '<year>_<4-digit-original-id>'   (stable citable key)
  SurveyYear  : int year
+ the 17 standard data columns carried over from the per-year FINAL files.

Rules:
  - No rows are removed. Cross-year duplicates (same route/point/band combo
    in different years) are valid distinct observations.
  - Where a year is missing a column (e.g. 2013 has no SurveyTime), the cell
    is left blank.
  - The per-year unique_id is preserved inside year_id, then dropped from
    the row (no duplicate ID columns).
"""

import os
import shutil
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# Paths
# ══════════════════════════════════════════════════════════════════════════════

SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT     = os.path.normpath(os.path.join(SCRIPT_DIR, "..", ".."))
OUTPUT_DIR    = os.path.join(REPO_ROOT, "Databases", "Database3", "Output")
ALL_YEARS_DIR = os.path.join(OUTPUT_DIR, "AllYears")
FINAL_PATH    = os.path.join(ALL_YEARS_DIR, "db3_ALL_YEARS_FINAL.xlsx")

DATABASE_NUMBER = 3   # this is Database 3


# ══════════════════════════════════════════════════════════════════════════════
# Column order in the combined Clean_Data sheet
# ══════════════════════════════════════════════════════════════════════════════

NEW_COLS = ["unique_id", "database", "year_id", "SurveyYear"]

DATA_COLS = [
    "SurveyDate", "SurveyTime", "WeatherCondition",
    "Route", "Latitude", "Longitude",
    "GroupNumber", "TotalObserved",
    "Observer", "ObserverEmail",
    "FlagCode", "FlagColor", "BandCombo",
    "Comments",
    "source_database", "source_file", "source_sheet",
]

FINAL_COL_ORDER = NEW_COLS + DATA_COLS


# ══════════════════════════════════════════════════════════════════════════════
# Read all per-year Clean_Data sheets
# ══════════════════════════════════════════════════════════════════════════════

def load_year(year: str) -> pd.DataFrame:
    """Read Clean_Data for one year. Return DataFrame with year_id added,
    original unique_id consumed, and data columns padded to the standard set."""
    path = os.path.join(OUTPUT_DIR, year, f"db3_{year}_FINAL.xlsx")
    if not os.path.exists(path):
        return None
    df = pd.read_excel(path, sheet_name="Clean_Data")

    # Build year_id from original per-year unique_id, padded to 4 digits
    if "unique_id" in df.columns:
        df["year_id"] = df["unique_id"].apply(lambda x: f"{year}_{int(x):04d}")
        df = df.drop(columns=["unique_id"])
    else:
        # Defensive fallback — should never happen
        df["year_id"] = [f"{year}_{i+1:04d}" for i in range(len(df))]

    df["SurveyYear"] = int(year)
    df["database"]   = DATABASE_NUMBER

    # Pad missing data columns with NaN so concat keeps a stable shape
    for col in DATA_COLS:
        if col not in df.columns:
            df[col] = pd.NA

    # Drop any internal pipeline columns that might have leaked through
    for col in ["_removal_reason", "_geo_warning", "_geo_correction"]:
        if col in df.columns:
            df = df.drop(columns=[col])

    return df


print(f"\n{'='*60}")
print("Combining all years of Database 3")
print(f"  Output dir: {ALL_YEARS_DIR}")
print()

# Ensure the AllYears folder exists
os.makedirs(ALL_YEARS_DIR, exist_ok=True)

year_dirs = sorted(
    d for d in os.listdir(OUTPUT_DIR)
    if d.isdigit() and os.path.isdir(os.path.join(OUTPUT_DIR, d))
)

year_frames = []
loaded_years = []
for year in year_dirs:
    df = load_year(year)
    if df is None or len(df) == 0:
        print(f"  {year}: SKIPPED (no FINAL file or empty)")
        continue
    year_frames.append(df)
    loaded_years.append(year)

    # Copy the per-year FINAL file into the AllYears folder
    src = os.path.join(OUTPUT_DIR, year, f"db3_{year}_FINAL.xlsx")
    dst = os.path.join(ALL_YEARS_DIR, f"db3_{year}_FINAL.xlsx")
    if os.path.exists(src):
        shutil.copy2(src, dst)

    print(f"  {year}: {len(df):>5} rows loaded, FINAL file copied")

print(f"\n  Total rows to combine: {sum(len(f) for f in year_frames)}")
print(f"  Years included:        {loaded_years}")

combined = pd.concat(year_frames, ignore_index=True, sort=False)

# Assign new global sequential unique_id (1..N) in year order
combined["unique_id"] = range(1, len(combined) + 1)

# Enforce final column order
combined = combined[FINAL_COL_ORDER]


# ══════════════════════════════════════════════════════════════════════════════
# Summary sheet
# ══════════════════════════════════════════════════════════════════════════════

def safe_int(x):
    try:    return int(x)
    except: return 0

def pct(num, den):
    return round(num / den * 100, 1) if den else 0.0

summary_rows = []
for year in loaded_years:
    sub = combined[combined["SurveyYear"] == int(year)]
    n_rows   = len(sub)
    n_routes = sub["Route"].nunique() if "Route" in sub.columns else 0
    total_pipl = sub["TotalObserved"].fillna(0).apply(safe_int).sum()
    banded   = sub[sub["BandCombo"].notna()]["TotalObserved"].fillna(0).apply(safe_int).sum()
    unbanded = sub[sub["BandCombo"].isna()]["TotalObserved"].fillna(0).apply(safe_int).sum()
    summary_rows.append({
        "Year":              int(year),
        "Routes Surveyed":   n_routes,
        "Total Rows":        n_rows,
        "Total PIPL":        total_pipl,
        "Banded":            banded,
        "Unbanded":          unbanded,
        "% Banded":          f"{pct(banded, total_pipl)}%",
    })

# Totals row across all years
all_pipl  = sum(r["Total PIPL"] for r in summary_rows)
all_band  = sum(r["Banded"]     for r in summary_rows)
all_unb   = sum(r["Unbanded"]   for r in summary_rows)
summary_rows.append({
    "Year":              "TOTAL",
    "Routes Surveyed":   combined["Route"].nunique(),
    "Total Rows":        len(combined),
    "Total PIPL":        all_pipl,
    "Banded":            all_band,
    "Unbanded":          all_unb,
    "% Banded":          f"{pct(all_band, all_pipl)}%",
})

summary_df = pd.DataFrame(summary_rows)


# ══════════════════════════════════════════════════════════════════════════════
# Write Excel
# ══════════════════════════════════════════════════════════════════════════════

with pd.ExcelWriter(FINAL_PATH, engine="openpyxl") as writer:
    combined.to_excel(writer,   sheet_name="All_Years_Combined", index=False)
    summary_df.to_excel(writer, sheet_name="Summary",            index=False)

    # One sheet per year — same column shape as the combined sheet, filtered.
    for year in loaded_years:
        year_df = combined[combined["SurveyYear"] == int(year)].copy()
        year_df.to_excel(writer, sheet_name=year, index=False)

    wb = writer.book
    header_font   = Font(bold=True, color="FFFFFF")
    combined_fill = PatternFill("solid", fgColor="2E7D32")
    summary_fill  = PatternFill("solid", fgColor="1565C0")
    year_fill     = PatternFill("solid", fgColor="6A1B9A")
    totals_fill   = PatternFill("solid", fgColor="E3F2FD")
    totals_font   = Font(bold=True, color="0D47A1")

    sheet_fills = {"All_Years_Combined": combined_fill, "Summary": summary_fill}
    for year in loaded_years:
        sheet_fills[year] = year_fill

    for sheet_name, fill in sheet_fills.items():
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")

        # Date format for SurveyDate on data sheets
        if sheet_name != "Summary":
            for col_cells in ws.iter_cols(min_row=1, max_row=1):
                if col_cells[0].value == "SurveyDate":
                    letter = get_column_letter(col_cells[0].column)
                    for cell in ws[letter][1:]:
                        cell.number_format = "MM/DD/YYYY"

        # Highlight TOTAL row in Summary
        if sheet_name == "Summary":
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value).strip().upper() == "TOTAL":
                    for c in row:
                        c.font = totals_font
                        c.fill = totals_fill

        # Auto-width columns (capped)
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 50)


print(f"\n{'─'*50}")
print(f"[DONE] AllYears folder: {ALL_YEARS_DIR}")
print(f"  Per-year FINAL files copied: {len(loaded_years)}")
print(f"  Combined workbook:           db3_ALL_YEARS_FINAL.xlsx")
print(f"    All_Years_Combined : {len(combined)} rows, {len(combined.columns)} columns")
print(f"    Summary            : {len(summary_df)} rows ({len(loaded_years)} years + 1 TOTAL)")
print(f"    Per-year sheets    : {', '.join(loaded_years)}")
print()
print("Quick stats:")
print(f"  Years included    : {loaded_years[0]} → {loaded_years[-1]}  ({len(loaded_years)} years)")
print(f"  Total PIPL counted: {all_pipl}")
print(f"  Banded            : {all_band}  ({pct(all_band, all_pipl)}%)")
print(f"  Unbanded          : {all_unb}  ({pct(all_unb, all_pipl)}%)")
print(f"  Unique routes     : {combined['Route'].nunique()}")
