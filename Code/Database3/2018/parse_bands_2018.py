"""
parse_bands_2018.py
-------------------
Pre-processing step for 2018 Winter Bird Survey band data.

Goal: for each {n}PIPLbands cell, determine HOW MANY individual birds'
information is present and number them clearly:
    "1) <raw text for bird 1>
     2) <raw text for bird 2>"

We do NOT need to interpret the band notation — just split and count.
If we can't determine the count at all, highlight pink for biologist review.

Input:  Databases/Database3Clean/Winter Birds 2018 Clean.xlsx  (34 routes)
Output: Databases/Database3BandReview/Winter Birds 2018 Band Review.xlsx
"""

import re
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

# ── Paths ────────────────────────────────────────────────────────────────────
ROOT  = Path(__file__).resolve().parents[3]          # …/PipingPlover/
CLEAN = ROOT / "Databases/Database3Clean/Winter Birds 2018 Clean.xlsx"
OUT   = ROOT / "Databases/Database3BandReview/Winter Birds 2018 Band Review.xlsx"

PINK = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

# ── No-band synonyms ─────────────────────────────────────────────────────────
_NONE_EXACT = {
    "none", "n/a", "na", "0", "no", "none.", "n.a.",
    "no bands", "no bands.", "none banded", "no banded birds",
    "no bands or flags", "no bands or flags.", "not banded",
    "both were unbanded.", "both unbanded", "unbanded",
}

def _is_no_band(text: str) -> bool:
    t = text.lower().strip().rstrip(".")
    return (t in {s.rstrip(".") for s in _NONE_EXACT}
            or t.startswith("no band")
            or t.startswith("not band")
            or t.startswith("both were unbanded")
            or t.startswith("both unbanded"))

def _is_unreadable(text: str) -> bool:
    """Bands present but illegible — count genuinely unknown → pink."""
    t = text.lower()
    return "none readable" in t or "not readable" in t or "unreadable" in t


# ── Trailing metadata noise ───────────────────────────────────────────────────
# Strip reporter/admin notes appended after real band text.
_TRAILING_NOISE = re.compile(
    r"[\-–]?\s*"
    r"(?:reported(?: to banders?| to great lakes(?: area)?(?: plover)?(?:\s+\w+)*)?|"
    r"REPORTED TO BANDERS?|"
    r"not reported to another site|"
    r"will be batch reported at end of season|"
    r"photos? (?:taken|available[^.]*)|"
    r"photo;[^)]*)"
    r"[.,]?\s*$",
    re.IGNORECASE,
)

# Leading "X banded." or "X birds banded." count prefix on the whole cell
_LEADING_COUNT = re.compile(r"^\d+\s+(?:banded|birds?\s+banded)[.,]?\s*", re.IGNORECASE)


def _clean(entry: str) -> str:
    """Strip trailing metadata noise from a single bird entry."""
    return _TRAILING_NOISE.sub("", entry).strip().strip(",").strip()


def _fmt(entries: list[str]) -> str:
    """Format a list of bird entries as '1) ... \\n2) ...'"""
    cleaned = [_clean(e) for e in entries if e and e.strip()]
    cleaned = [e for e in cleaned if e]
    return "\n".join(f"{i+1}) {e}" for i, e in enumerate(cleaned))


# ── Splitters ─────────────────────────────────────────────────────────────────

def _split_depth0(text: str, seps=(",", ";")) -> list[str]:
    """
    Split on ', ' or '; ' only when not inside parentheses (depth 0).
    Preserves content inside parens like (O64 or O6Y) intact.
    """
    parts, buf, depth = [], [], 0
    i = 0
    while i < len(text):
        c = text[i]
        if c == "(":
            depth += 1; buf.append(c)
        elif c == ")":
            depth -= 1; buf.append(c)
        elif depth == 0 and c in seps and i + 1 < len(text) and text[i + 1] == " ":
            parts.append("".join(buf).strip())
            buf = []; i += 2; continue
        else:
            buf.append(c)
        i += 1
    if buf:
        parts.append("".join(buf).strip())
    return [p for p in parts if p]


# Patterns where the observer already numbered birds
_NUMBERED_RE = [
    re.compile(r"(?:^|\s)(\d+)\)\s*"),           # 1) 2) 3)
    re.compile(r"(?:^|\n)\s*(\d+)\.\s+"),         # 1. 2. 3.
    re.compile(r"PP(\d+):\s*", re.IGNORECASE),    # PP1: PP2:
]

def _try_numbered(text: str) -> list[str] | None:
    for pat in _NUMBERED_RE:
        if pat.search(text):
            parts = re.split(pat.pattern, text, flags=pat.flags)
            # split with capture group gives [pre, g1, content, g1, content ...]
            # drop empties and bare digit captures
            entries = [p.strip() for p in parts
                       if p and p.strip() and not re.fullmatch(r"\d+", p.strip())]
            if entries:
                return entries
    return None


def _try_newlines(text: str) -> list[str] | None:
    """
    Newline-separated birds. Lenient: just needs multiple non-empty lines.
    Each line is one bird's info regardless of notation style.
    """
    if "\n" not in text:
        return None
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    if len(lines) >= 2:
        return lines
    return None


def _try_slash_slash(text: str) -> list[str]:
    """
    Standard UL//LL:UR//LR notation — split on depth-0 ', ' or '; '.
    Strip leading 'N: ' count prefix first.
    """
    body = re.sub(r"^\d+:\s*", "", text)
    return _split_depth0(body)


def _try_semicolons(text: str) -> list[str] | None:
    """
    Some observers separate birds with '; ' without using // notation.
    Only use this if we get ≥2 parts that each look like band info
    (contain ':', ',', or a known leg keyword).
    """
    parts = [p.strip() for p in text.split(";") if p.strip()]
    if len(parts) < 2:
        return None
    def looks_like_band(p):
        p = p.lower()
        return any(k in p for k in (":", ",", "left", "right", "ll=", "rl=", "//"))
    if sum(looks_like_band(p) for p in parts) >= 2:
        return parts
    return None


# ── Main parser ───────────────────────────────────────────────────────────────

def parse_band_cell(raw) -> tuple[str | None, bool]:
    """
    Returns (parsed_text, needs_review).

    parsed_text is always "1) ...\n2) ..." format showing one entry per bird.
    needs_review=True (→ pink) ONLY when we cannot determine the count at all.

    We do NOT need to understand the notation — just count and separate birds.
    When in doubt, assume 1 bird and keep the raw text as-is.
    """
    if raw is None:
        return None, False
    text = str(raw).strip()
    if not text:
        return None, False

    # 1. No banded birds
    if _is_no_band(text):
        return "No banded birds", False

    # 2. Bands present but illegible → count unknown → pink
    if _is_unreadable(text):
        return text, True

    # Strip leading "X banded." prefix before further processing
    text = _LEADING_COUNT.sub("", text).strip()

    # 3. Observer already numbered birds (1), 1., PP1:)
    numbered = _try_numbered(text)
    if numbered:
        return _fmt(numbered), False

    # 4. Newline-separated — each line = one bird
    newlines = _try_newlines(text)
    if newlines:
        return _fmt(newlines), False

    # 5. Standard // notation — depth-0 comma/semicolon split
    if "//" in text:
        birds = _try_slash_slash(text)
        if birds:
            return _fmt(birds), False

    # 6. Semicolon-separated without // (some observers use ; as bird separator)
    semis = _try_semicolons(text)
    if semis:
        return _fmt(semis), False

    # 7. Default: treat entire cell as 1 bird's info.
    #    Keep raw text — we don't need to understand it, just know count = 1.
    return f"1) {_clean(text)}", False


# ── Identify PIPLbands columns ───────────────────────────────────────────────

def find_pipl_band_cols(headers: list) -> list[int]:
    """Return 0-based column indices for all {n}PIPLbands columns."""
    return [
        i for i, h in enumerate(headers)
        if h and re.fullmatch(r"\d+PIPL.?ands", str(h), re.IGNORECASE)
    ]


# ── Build output workbook ─────────────────────────────────────────────────────

def run():
    print(f"Reading: {CLEAN}")
    wb_in = openpyxl.load_workbook(CLEAN)
    ws_in = wb_in["Sheet1"]

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet1"

    headers = [cell.value for cell in ws_in[1]]
    band_cols = set(find_pipl_band_cols(headers))

    print(f"Rows: {ws_in.max_row - 1} data rows  |  PIPLbands columns: {len(band_cols)}")

    stats = {"total": 0, "no_band": 0, "parsed": 0, "pink": 0}

    for row_idx, row in enumerate(ws_in.iter_rows(), start=1):
        out_row = []
        results = {}   # col_idx → (parsed, needs_review)

        for col_idx, cell in enumerate(row):
            if row_idx == 1 or col_idx not in band_cols:
                out_row.append(cell.value)
            else:
                parsed, needs_review = parse_band_cell(cell.value)
                results[col_idx] = (parsed, needs_review)
                out_row.append(parsed)

        ws_out.append(out_row)

        if row_idx > 1:
            for col_idx, (parsed, needs_review) in results.items():
                raw_val = ws_in.cell(row=row_idx, column=col_idx + 1).value
                if raw_val is not None and str(raw_val).strip():
                    stats["total"] += 1
                    if parsed == "No banded birds":
                        stats["no_band"] += 1
                    elif needs_review:
                        stats["pink"] += 1
                    else:
                        stats["parsed"] += 1
                if needs_review:
                    ws_out.cell(row=row_idx, column=col_idx + 1).fill = PINK

    ws_out.freeze_panes = "A2"
    wb_out.save(OUT)

    print(f"Saved:   {OUT}")
    print()
    print("── Stats ───────────────────────────────────────────────")
    print(f"  Non-empty band cells:  {stats['total']}")
    print(f"  → No banded birds:     {stats['no_band']}")
    print(f"  → Numbered (≥1 bird):  {stats['parsed']}")
    print(f"  → Needs review (pink): {stats['pink']}  ← count genuinely unknown")


if __name__ == "__main__":
    run()
