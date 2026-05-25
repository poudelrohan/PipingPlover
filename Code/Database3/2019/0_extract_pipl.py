"""
Step 0: Extract PIPL Data — 2019
──────────────────────────────────
Reads the 2019 Band Review file (wide format, one row per route) and melts
it into one row per bird per point, applying Option A band expansion.

Band data source priority per (route, point):
  1. Biologist correction  — Biologist_Band_Review_Completed_2018-2019.xlsx col H
  2. Our interpretation    — same file col F (already structured 1)/2)/3))
  3. Raw Band Review text  — fallback if the review file has no entry for that point

2019-specific handling:
  ① Route names have ", County=X" suffix → stripped before output
  ② Point 1 has no longitude in Band Review / Clean file → read from raw file col 63
  ③ New Smyrna Beach has two rows (10:30 duplicate dropped, 09:15 kept)
  ④ All 19 points present
  ⑤ No observer email column in the 2019 Google Form
  ⑥ GPS fallback: 2020–2024 data used if a point still has no GPS

Output: db3_2019_extracted.xlsx  (in Output/2019/)
"""

import pandas as pd
import re
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))
from database3_config import config, get_output_folder, get_filename

script_dir    = os.path.dirname(os.path.abspath(__file__))
output_folder = get_output_folder(script_dir)
year          = config["year"]

review_dir  = os.path.normpath(os.path.join(script_dir, config["input_folder"]))
review_path = os.path.join(review_dir, config["file"])
raw_dir     = os.path.normpath(os.path.join(script_dir, config["raw_folder"]))
raw_path    = os.path.join(raw_dir, config["raw_file"])
output_path  = os.path.join(output_folder, get_filename("extracted"))
removed_path = os.path.join(output_folder, get_filename("removed_gps"))

points = config["points"]
cols   = config["columns"]


# ══════════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════════

def normalize_route(s):
    s = str(s).lower().strip()
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def strip_county(s):
    """Remove ', County=...' suffix from a route name."""
    return re.sub(r',\s*County=.*$', '', str(s), flags=re.IGNORECASE).strip()

def parse_coord(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        f = float(str(val).strip())
        return f if f != 0.0 else None
    except (ValueError, TypeError):
        return None

def combine_weather(temp, wind, rain):
    parts = []
    for v in (temp, wind, rain):
        s = "" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v).strip()
        if s and s.lower() not in ("nan", "none", ""):
            parts.append(s)
    return ", ".join(parts) if parts else None

def parse_band_entries(text):
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return []
    text = str(text).strip()
    if not text or text.lower() == "no banded birds":
        return []
    # Normalise inline numbering: "1) bird1 2) bird2" → "1) bird1\n2) bird2"
    # so the newline-based splitter below handles both layouts.
    text = re.sub(r'(?<!\n)(\s+)(\d+\)\s*)', r'\n\2', text)
    entries = []
    for line in text.split("\n"):
        line = line.strip()
        m = re.match(r'^\d+\)\s*', line)
        if m:
            entry = line[m.end():].strip()
            if entry:
                entries.append(entry)
    # fallback: whole cell = 1 bird
    if not entries and text:
        entries = [text]
    return entries


# ══════════════════════════════════════════════════════════════════════════════
# Biologist Corrections
# ══════════════════════════════════════════════════════════════════════════════

def _normalize_correction_fmt(text: str) -> str:
    """
    Fix common formatting issues biologists introduce when typing corrections:
      • '2). text'  →  '2) text'   (extra dot after closing paren)
      • ' 4. text'  →  '\\n4) text'  (period-style inline bullets → paren-style)
    """
    # Strip extra dot right after a closing paren: "2)." → "2)"
    text = re.sub(r'(\d+\))\.\s*', r'\1 ', text)
    # Convert inline " N. " period-bullets to "\nN) " (e.g. " 4. x,B/R:" → "\n4) x,B/R:")
    text = re.sub(r'\s+(\d+)\.\s+', lambda m: f'\n{m.group(1)}) ', text)
    return text.strip()


def _resolve_band_text(correction: str, interpretation: str):
    """
    Decide what band text to use given the biologist's correction and our interpretation.

    Returns:
      None  → skip this (route, point) entirely (e.g. 'remove this row')
      ''    → treat as no banded birds (e.g. 'unreadable, make unbanded')
      str   → the band text to parse (corrected, or confirmed interpretation)
    """
    if not correction:
        return interpretation or None

    t = correction.lower().strip()

    # "Remove / delete" → skip row
    if any(p in t for p in ("remove", "delete")):
        return None

    # "Unreadable / unbanded" instruction
    if ("unbanded" in t or "unreadable" in t) and "1)" not in t and "/" not in t:
        return ""

    # "This is good" / confirmation → our interpretation is correct
    if ("this is good" in t or "that is good" in t) and "/" not in t and "1)" not in t:
        return interpretation

    # Actual band data — normalise formatting then return
    return _normalize_correction_fmt(correction)


def build_biologist_corrections():
    """
    Read the completed biologist review file (2019 sheet) and build:
        (normalized_route, point_N) → final_band_text
    """
    review_path_bio = (
        Path(script_dir).parents[2]
        / "Databases" / "Database3BiologistReview"
        / "Biologist_Band_Review_Completed_2018-2019.xlsx"
    )
    if not review_path_bio.exists():
        print("  [INFO] No biologist review file found — using raw Band Review data")
        return {}

    import openpyxl
    wb  = openpyxl.load_workbook(review_path_bio)
    ws  = wb["2019"]
    corrections = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        route      = str(row[0] or "").strip()
        point      = row[1]
        interp     = str(row[5] or "").strip()   # col F — our interpretation
        correction = str(row[7] or "").strip()   # col H — biologist correction

        if not route or not point:
            continue

        text = _resolve_band_text(correction, interp)
        if text is None:
            continue   # skip row (delete instruction)

        key = (normalize_route(strip_county(route)), int(point))
        # Later rows with same key overwrite earlier ones
        # (handles New Smyrna: row 18 = skip, row 19 = FN(U9) ✓)
        corrections[key] = text

    n_band     = sum(1 for v in corrections.values() if v)
    n_unbanded = sum(1 for v in corrections.values() if v == "")
    print(f"  Biologist corrections loaded: {n_band} band entries, "
          f"{n_unbanded} points overridden to unbanded")
    return corrections


# ══════════════════════════════════════════════════════════════════════════════
# ② Read Point 1 longitude from raw file
# ══════════════════════════════════════════════════════════════════════════════

def build_pt1_lon_lookup():
    """
    The Band Review file has no Point 1 longitude column (the raw Google Form
    column '1 Longitude...' was excluded during cleaning).
    Read it directly from the raw 2019 Excel (col 63) and build a dict:
        normalized_route → longitude
    """
    lookup = {}
    import openpyxl
    wb = openpyxl.load_workbook(raw_path)
    ws = wb[config["raw_sheet"]]
    headers = [cell.value for cell in ws[1]]

    # Find the '1 Longitude' column (col 63 in raw file)
    lon_col = next(
        (i for i, h in enumerate(headers)
         if h and re.search(r'^1\s+Longitude', str(h), re.IGNORECASE)), None
    )
    route_col = next(
        (i for i, h in enumerate(headers)
         if h and 'transect' in str(h).lower()), None
    )
    if route_col is None:
        route_col = 0

    if lon_col is None:
        print("  [WARNING] Could not find '1 Longitude' column in raw file")
        return lookup

    print(f"  Raw file Point 1 lon col: {lon_col} = '{str(headers[lon_col])[:60]}'")

    for row in ws.iter_rows(min_row=2, values_only=True):
        route_raw = row[route_col]
        lon_val   = parse_coord(row[lon_col])
        if route_raw and lon_val:
            key = normalize_route(strip_county(str(route_raw)))
            if key not in lookup:
                lookup[key] = lon_val

    print(f"  Point 1 longitude lookup: {len(lookup)} routes")
    return lookup


# ══════════════════════════════════════════════════════════════════════════════
# ⑥ GPS fallback from 2020–2024
# ══════════════════════════════════════════════════════════════════════════════

def build_gps_fallback():
    """
    Scan 2020–2024 Focal Observations clean sheets and build:
        (normalized_route, point_N) → (lat, lon, source_year)
    Used only if a point still has no GPS after the raw-file fix.
    """
    lookup = {}
    root   = Path(script_dir).parents[2] / "Databases" / "Database3Clean"

    for yr in range(2020, 2025):
        fpath = root / f"Winter Birds {yr} Clean.xlsx"
        if not fpath.exists():
            continue
        try:
            wb_tmp = pd.ExcelFile(fpath)
        except Exception:
            continue
        if "Focal Observations" not in wb_tmp.sheet_names:
            continue
        df_tmp = wb_tmp.parse("Focal Observations", header=0)
        route_col_name = df_tmp.columns[0]

        for n in range(1, 20):
            lat_col = next((c for c in df_tmp.columns
                            if re.search(rf'Lat.*\bpoint\s+{n}\b', str(c), re.IGNORECASE)
                            or re.search(rf'\bpoint\s+{n}\b.*Lat', str(c), re.IGNORECASE)), None)
            lon_col = next((c for c in df_tmp.columns
                            if re.search(rf'Long.*\bpoint\s+{n}\b', str(c), re.IGNORECASE)
                            or re.search(rf'\bpoint\s+{n}\b.*Long', str(c), re.IGNORECASE)), None)
            if lat_col is None or lon_col is None:
                continue
            for _, r in df_tmp.iterrows():
                lat = parse_coord(r.get(lat_col))
                lon = parse_coord(r.get(lon_col))
                route_val = r.get(route_col_name)
                if route_val and lat and lon:
                    key = (normalize_route(strip_county(str(route_val))), int(n))
                    if key not in lookup:
                        lookup[key] = (lat, lon, yr)

    print(f"  GPS fallback lookup: {len(lookup)} (route, point) pairs from 2020–2024")
    return lookup


# ══════════════════════════════════════════════════════════════════════════════
# Map column indices from Band Review headers
# ══════════════════════════════════════════════════════════════════════════════

def build_point_col_map(headers):
    """
    Returns dict: point_N → {lat, long, pipl, band} column indices.
    Handles the long column name format of the 2019 Google Form.
    """
    point_map = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        h_str = str(h)

        # Lat: "Point N Latitude..."
        m = re.search(r'[Pp]oint\s+(\d+)\s+Lat', h_str)
        if m:
            n = int(m.group(1))
            point_map.setdefault(n, {})["lat"] = i

        # Long: "Point N Longitude..."
        m = re.search(r'[Pp]oint\s+(\d+)\s+Long', h_str)
        if m:
            n = int(m.group(1))
            point_map.setdefault(n, {})["long"] = i

        # PIPL count: "N Number of Piping Plovers..."
        m = re.match(r'^(\d+)\s+Number of Piping', h_str, re.IGNORECASE)
        if m:
            n = int(m.group(1))
            point_map.setdefault(n, {})["pipl"] = i

        # Band: "N Band/Flag Codes for Piping Plovers..."
        m = re.match(r'^(\d+)\s+Band/Flag Codes for Piping', h_str, re.IGNORECASE)
        if m:
            n = int(m.group(1))
            point_map.setdefault(n, {})["band"] = i

    return point_map


# ══════════════════════════════════════════════════════════════════════════════
# Main extraction
# ══════════════════════════════════════════════════════════════════════════════

print(f"\n{'='*60}")
print(f"Step 0 — 2019 PIPL extraction")
print(f"  Review: {review_path}")
print(f"  Raw:    {raw_path}")
print()

df_raw = pd.read_excel(review_path, sheet_name=config["sheet"], header=config["header_row"])
print(f"  Loaded {len(df_raw)} route rows from Band Review")

# ── Load biologist corrections ─────────────────────────────────────────────────
bio_corrections = build_biologist_corrections()

# ── ② Build Point 1 longitude lookup from raw file ────────────────────────────
pt1_lon_lookup = build_pt1_lon_lookup()

# ── ⑥ Build GPS fallback from 2020–2024 ──────────────────────────────────────
gps_fallback = build_gps_fallback()

# ── ③ Drop the incomplete New Smyrna Beach duplicate (10:30, no GPS) ─────────
route_col_idx = cols["route"]   # integer index
nsb_mask = df_raw.iloc[:, route_col_idx].astype(str).str.contains("New Smyrna", case=False, na=False)
nsb_rows  = df_raw[nsb_mask]

# Rows removed before the main loop — stored here and added to out_rows
# with _removal_reason so they flow through Steps 1-5 and appear in Removed_Rows.
pre_removed_rows = []

if len(nsb_rows) == 2:
    # Identify the row with no Point 1 latitude (the incomplete one)
    headers_list = list(df_raw.columns)
    pt_map_tmp   = build_point_col_map(headers_list)
    lat1_col     = headers_list[pt_map_tmp.get(1, {}).get("lat", -1)] if 1 in pt_map_tmp else None

    if lat1_col:
        no_gps_mask = nsb_mask & df_raw[lat1_col].isna()
    else:
        time_col_name = cols["date"] if "time" not in cols else cols["time"]
        no_gps_mask   = nsb_mask & (df_raw.get(cols["time"], pd.Series()).astype(str).str.contains("10:30"))

    n_dropped = no_gps_mask.sum()
    if n_dropped > 0:
        dropped_nsb = df_raw[no_gps_mask].iloc[0]
        pre_removed_rows.append({
            "SurveyDate":      dropped_nsb.get(cols["date"]),
            "Route":           strip_county(str(dropped_nsb.iloc[route_col_idx])),
            "GroupNumber":     1,
            "TotalObserved":   None,
            "BandCombo":       None,
            "_removal_reason": "Duplicate New Smyrna Beach submission: same observer/date/bird (U9) as 09:15 row but missing GPS — dropped in favour of the complete 09:15 record",
        })
        df_raw = df_raw[~no_gps_mask].reset_index(drop=True)
        print(f"  [FIX ③] Dropped incomplete New Smyrna Beach duplicate (10:30, no GPS). {n_dropped} row(s) removed.")
else:
    print(f"  [INFO] New Smyrna Beach rows found: {len(nsb_rows)} (expected 2 for dedup fix)")

print(f"  Working rows after dedup: {len(df_raw)}")

# ── Map point columns ──────────────────────────────────────────────────────────
headers_list = list(df_raw.columns)
point_cols   = build_point_col_map(headers_list)

# ── Find metadata columns ──────────────────────────────────────────────────────
def find_col(pattern):
    """Find a column index by exact name or substring."""
    if isinstance(pattern, int):
        return pattern
    for i, h in enumerate(headers_list):
        if str(h) == pattern or (pattern and pattern in str(h)):
            return i
    return None

date_col     = find_col(cols["date"])
time_col     = find_col(cols["time"])
temp_col     = find_col(cols["temp"])
wind_col     = find_col(cols["wind"])
rain_col     = find_col(cols["rain"])
observer_col = find_col(cols["observer"])
comments_col = find_col(cols["comments"])

# ── Melt wide → long with Option A expansion ──────────────────────────────────
out_rows = []

stats = {
    "routes":                 len(df_raw),
    "points_skipped_no_pipl": 0,
    "points_dropped_no_gps":  0,
    "pt1_lon_from_raw":       0,
    "points_gps_fallback":    0,
    "rows_banded":            0,
    "rows_unbanded":          0,
    "total_pipl":             0,
}

for _, route_row in df_raw.iterrows():

    route_raw = str(route_row.iloc[route_col_idx] or "").strip()
    route     = strip_county(route_raw)
    route_key = normalize_route(route)

    date_val = route_row.iloc[date_col] if date_col is not None else None
    time_val = route_row.iloc[time_col] if time_col is not None else None
    weather  = combine_weather(
        route_row.iloc[temp_col] if temp_col is not None else None,
        route_row.iloc[wind_col] if wind_col is not None else None,
        route_row.iloc[rain_col] if rain_col is not None else None,
    )
    observer = route_row.iloc[observer_col] if observer_col is not None else None
    comments_raw = route_row.iloc[comments_col] if comments_col is not None else None
    comments = str(comments_raw).strip() if (
        comments_raw is not None and not (isinstance(comments_raw, float) and pd.isna(comments_raw))
        and str(comments_raw).strip().lower() not in ("nan", "none", "")
    ) else None

    if pd.notna(date_val):
        try:
            date_val = pd.to_datetime(date_val)
        except Exception:
            pass

    for n in points:
        pc = point_cols.get(n, {})

        # PIPL count
        pipl_raw = route_row.iloc[pc["pipl"]] if "pipl" in pc else None
        try:
            pipl_count = int(float(pipl_raw)) if pipl_raw is not None and not (
                isinstance(pipl_raw, float) and pd.isna(pipl_raw)) else 0
        except (ValueError, TypeError):
            pipl_count = 0

        if pipl_count == 0:
            stats["points_skipped_no_pipl"] += 1
            continue

        # Latitude
        lat = parse_coord(route_row.iloc[pc["lat"]]) if "lat" in pc else None

        # Longitude — normal path
        lon = parse_coord(route_row.iloc[pc["long"]]) if "long" in pc else None

        gps_note = None

        # ② Point 1 longitude: fill from raw file lookup
        if n == 1 and lon is None:
            lon = pt1_lon_lookup.get(route_key)
            if lon is not None:
                stats["pt1_lon_from_raw"] += 1
                gps_note = "Point 1 longitude read from raw 2019 file (missing from Band Review)"

        # ⑥ Any remaining null GPS → try 2020–2024 fallback
        if lat is None or lon is None:
            fb_key = (route_key, n)
            if fb_key in gps_fallback:
                fb_lat, fb_lon, fb_yr = gps_fallback[fb_key]
                if lat is None:
                    lat = fb_lat
                if lon is None:
                    lon = fb_lon
                stats["points_gps_fallback"] += 1
                note = f"GPS borrowed from {fb_yr} — point {n} had no coordinates in 2019"
                gps_note = (gps_note + " | " + note) if gps_note else note

        if lat is None or lon is None:
            stats["points_dropped_no_gps"] += 1
            out_rows.append({
                "SurveyDate":      date_val,
                "Route":           route,
                "GroupNumber":     n,
                "TotalObserved":   pipl_count,
                "BandCombo":       None,
                "_removal_reason": f"No GPS for point {n} — not found in raw file or any other year",
            })
            print(f"  [DROP] {route} pt {n}: PIPL={pipl_count}, no GPS in any source")
            continue

        # Band entries — use biologist correction if available, else raw Band Review text
        band_raw = route_row.iloc[pc["band"]] if "band" in pc else None
        bio_key  = (route_key, n)
        if bio_key in bio_corrections:
            band_raw = bio_corrections[bio_key]   # "" = unbanded; string = corrected combo
        band_list = parse_band_entries(band_raw)
        n_banded  = len(band_list)
        remainder = pipl_count - n_banded

        if remainder < 0:
            print(f"  [WARNING] {route} pt {n}: {n_banded} banded entries > PIPL count "
                  f"({pipl_count}). Keeping all band rows — flag for biologist review.")
            remainder = 0

        base = {
            "SurveyDate":       date_val,
            "SurveyTime":       str(time_val).strip() if pd.notna(time_val) else None,
            "WeatherCondition": weather,
            "Route":            route,
            "Latitude":         lat,
            "Longitude":        lon,
            "GroupNumber":      n,
            "Observer":         str(observer).strip() if pd.notna(observer) else None,
            "ObserverEmail":    None,       # not in 2019 form
            "FlagCode":         None,
            "FlagColor":        None,
            "Comments":         comments,   # original observer text only — no pipeline notes
        }

        # Banded rows
        for band_text in band_list:
            row = base.copy()
            row["TotalObserved"] = 1
            row["BandCombo"]     = band_text
            out_rows.append(row)
            stats["rows_banded"] += 1

        # Unbanded remainder
        if remainder > 0:
            row = base.copy()
            row["TotalObserved"] = remainder
            row["BandCombo"]     = None
            out_rows.append(row)
            stats["rows_unbanded"] += 1
        elif n_banded == 0:
            row = base.copy()
            row["TotalObserved"] = pipl_count
            row["BandCombo"]     = None
            out_rows.append(row)
            stats["rows_unbanded"] += 1

        stats["total_pipl"] += pipl_count

# ── Save ──────────────────────────────────────────────────────────────────────
# Pre-pipeline removed rows (New Smyrna duplicate) are appended to out_rows
# with _removal_reason set. Step 3 recognises the flag and routes them to
# the pipeline's Removed_Rows sheet automatically.
all_rows = out_rows + pre_removed_rows
df_out   = pd.DataFrame(all_rows)

df_out.to_excel(output_path, index=False)

print(f"\n{'─'*50}")
print(f"[DONE] Step 0 — 2019")
print(f"  Routes processed:             {stats['routes']}")
print(f"  Points skipped (0 PIPL):      {stats['points_skipped_no_pipl']}")
print(f"  Pt1 lon read from raw file:   {stats['pt1_lon_from_raw']}")
print(f"  Points GPS from fallback:     {stats['points_gps_fallback']}")
print(f"  Points dropped (no GPS):      {stats['points_dropped_no_gps']}")
print(f"  Output rows — banded:         {stats['rows_banded']}")
print(f"  Output rows — unbanded:       {stats['rows_unbanded']}")
print(f"  Total output rows:            {len(df_out)}")
print(f"  Total PIPL:                   {stats['total_pipl']}")
print(f"  Extracted: {output_path}")
print(f"  (Pre-pipeline removed rows included in extracted file with _removal_reason set)")
