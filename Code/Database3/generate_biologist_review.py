"""
generate_biologist_review.py
────────────────────────────
Reads raw clean files (Database3Clean/) for the requested year(s),
interprets every non-trivial PIPLbands cell using BAND_LOGIC.md rules,
and writes a single Excel workbook for biologist review:

  Sheet 1 — How to Fill This      (instructions)
  Sheet 2 — 2018                  (one sheet per year)

Biologist workflow:
  • Leave "Correction" blank if our interpretation is correct.
  • Write the correct version in "Correction" only if something is wrong.

Usage:
    python3 generate_biologist_review.py
"""

import re
import os
import sys
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment,
                              Border, Side)
from openpyxl.utils import get_column_letter

ROOT    = Path(__file__).resolve().parents[2]
CLEAN   = ROOT / "Databases" / "Database3Clean"
OUT_DIR = ROOT / "Databases" / "Database3BiologistReview"
OUT_DIR.mkdir(exist_ok=True)
OUT     = OUT_DIR / "Biologist_Band_Review_2018-2024.xlsx"

# ── Colours ───────────────────────────────────────────────────────────────────
C_HEADER   = "1B4F72"   # dark blue  — header row
C_OK       = "FFFFFF"   # white      — confident interpretation
C_REVIEW   = "FFF9C4"   # pale yellow— ambiguous, please confirm
C_CONFLICT = "FFE0B2"   # pale orange— count conflict / data error
C_INST_HDR = "1B4F72"   # same dark blue for instructions header
C_INST_BG  = "EBF5FB"   # light blue for instruction body

# ── No-band synonyms (skip these) ─────────────────────────────────────────────
_NO_BAND = {
    "none","n/a","na","0","no","no bands","no bands seen","none banded",
    "not banded","both were unbanded","all unbanded","all xx","xx","x:x",
    "no bands or flags","none visible","n.a.","none.","no band",
    "no banded birds",
}
def is_no_band(text: str) -> bool:
    t = text.lower().strip().rstrip(".")
    return (t in _NO_BAND
            or t.startswith("no band")
            or t.startswith("not band")
            or t.startswith("both were unbanded")
            or t.startswith("all unbanded")
            or t.startswith("all xx"))

# ── Unreadable synonyms ────────────────────────────────────────────────────────
def is_unreadable(text: str) -> bool:
    t = text.lower()
    return any(p in t for p in [
        "none readable","not readable","unreadable",
        "too distant","unable to read","could not read",
        "partial resight","incomplete band combo",
    ])

# ── Strip trailing metadata noise ─────────────────────────────────────────────
_NOISE = re.compile(
    r"[\-–]?\s*"
    r"(?:reported[^.]{0,60}?(?:banders?|database|coordinator|researcher|plover@\S+|audubon|fwc|vt)|"
    r"REPORTED TO BANDERS?|"
    r"not reported to another site|"
    r"will (?:report|be batch reported)[^.]*|"
    r"photos? (?:taken|available[^.]*)|"
    r"banded (?:as (?:a chick|an adult)|in [^.]+)|"
    r"continuing bird[^.]*|"
    r"this bird has been sighted[^.]*)"
    r"[.,]?\s*$",
    re.IGNORECASE,
)
def strip_noise(s: str) -> str:
    return _NOISE.sub("", s).strip().strip(",").strip()

# ── Strip count prefix ────────────────────────────────────────────────────────
# Handles:  "5 banded PIPL seen, ..."
#           "4 Unbanded, 1 Banded - ..."   ← Little Talbot style
_COUNT_PFX = re.compile(
    r"^\d+\s+(?:banded\s+)?(?:pipl\s+)?(?:banded\s+)?(?:birds?\s+)?(?:banded\s+)?(?:seen[,.]?\s*|:?\s*)",
    re.IGNORECASE,
)
_UNBANDED_BANDED_PFX = re.compile(
    r"^\d+\s+unbanded[,\s]+\d+\s+banded\s*[-–]\s*",
    re.IGNORECASE,
)
def strip_count_prefix(s: str) -> str:
    s = _UNBANDED_BANDED_PFX.sub("", s).strip()
    s = _COUNT_PFX.sub("", s).strip()
    return s

# ── Depth-0 split on comma/semicolon ──────────────────────────────────────────
def split_depth0(text: str, seps=(",", ";")) -> list:
    parts, buf, depth = [], [], 0
    i = 0
    while i < len(text):
        c = text[i]
        if c == "(":   depth += 1; buf.append(c)
        elif c == ")": depth -= 1; buf.append(c)
        elif depth == 0 and c in seps and i+1 < len(text) and text[i+1] == " ":
            parts.append("".join(buf).strip()); buf = []; i += 2; continue
        else: buf.append(c)
        i += 1
    if buf: parts.append("".join(buf).strip())
    return [p for p in parts if p]

# ── Numbered-entry patterns ────────────────────────────────────────────────────
_NUMBERED = [
    re.compile(r"(?:^|\s)(\d+)\)\s*"),
    re.compile(r"(?:^|\n)\s*(\d+)\.\s+"),
    re.compile(r"PP(\d+):\s*",  re.IGNORECASE),
    re.compile(r"PIPL\s*([A-Z]):\s*", re.IGNORECASE),
]
def try_numbered(text: str):
    for pat in _NUMBERED:
        if pat.search(text):
            pieces = re.split(pat.pattern, text, flags=pat.flags)
            parts = [p.strip() for p in pieces
                     if p and p.strip() and not re.fullmatch(r"[\dA-Z]", p.strip())]
            if parts:
                return parts
    return None

# ── Normalise inline numbering: "1) a 2) b" → "1) a\n2) b" ───────────────────
def normalise_inline(text: str) -> str:
    return re.sub(r"(?<!\n)(\s+)(\d+\)\s*)", r"\n\2", text)

# ── Main interpreter ───────────────────────────────────────────────────────────
def interpret(raw, pipl_count: int):
    """
    Returns (entries: list[str], status: str, notes: str)
      entries  — individual bird band strings after splitting
      status   — "ok" | "review" | "conflict"
      notes    — explanation shown to biologist
    """
    if raw is None:
        return [], "ok", ""
    text = str(raw).strip()
    if not text:
        return [], "ok", ""

    if is_no_band(text):
        return [], "ok", ""         # skipped — no banded birds

    if is_unreadable(text):
        return [text], "review", "Bands present but could not be read — please confirm count"

    # Strip leading count prefix
    text = strip_count_prefix(text)
    if not text:
        return [], "ok", ""

    # Normalise inline numbering before splitting
    text = normalise_inline(text)

    # Try numbered entries
    entries = try_numbered(text)

    # Try newline split
    if not entries and "\n" in text:
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        if len(lines) >= 2:
            entries = lines

    # Try // with depth-0 comma/semicolon
    if not entries and "//" in text:
        body = re.sub(r"^\d+:\s*", "", text)
        parts = split_depth0(body)
        if parts:
            entries = parts

    # Try " and " separator
    if not entries and " and " in text.lower():
        parts = re.split(r"\s+and\s+", text, flags=re.IGNORECASE)
        parts = [p.strip() for p in parts
                 if p.strip() and not re.fullmatch(r"\d+\s+unbanded", p.strip(), re.IGNORECASE)]
        if len(parts) >= 2:
            entries = parts

    # Try multi-space separator
    if not entries and re.search(r"   +", text):
        parts = [p.strip() for p in re.split(r"   +", text) if p.strip()]
        if len(parts) >= 2:
            entries = parts

    # Try semicolon separator — flag as review since count may be ambiguous
    semi_used = False
    if not entries and ";" in text:
        parts = [p.strip() for p in text.split(";") if p.strip()]
        if len(parts) >= 2:
            entries = parts
            semi_used = True

    # Default: whole cell = 1 bird
    if not entries:
        entries = [text]

    # Strip noise from each entry; remove pure "X:X" unbanded entries
    cleaned = []
    unbanded_from_xx = 0
    for e in entries:
        e = strip_noise(e).strip()
        if not e:
            continue
        if re.fullmatch(r"[Xx][:/\\][Xx]|[Xx]{2}", e):
            unbanded_from_xx += 1   # X:X = explicitly unbanded, don't make a banded row
        else:
            cleaned.append(e)

    entries = cleaned
    n_banded = len(entries)
    notes_parts = []

    if unbanded_from_xx:
        notes_parts.append(
            f"{unbanded_from_xx} entry/entries were X:X (unbanded) — removed from banded list; "
            f"counted as part of unbanded remainder"
        )

    if semi_used:
        notes_parts.append(
            f"Semicolons used as bird separator — we read {n_banded} bird(s). "
            f"Please confirm this is correct."
        )

    # Count validation
    if pipl_count == 0:
        return [], "ok", "PIPL count is 0 — band data ignored"

    if n_banded > pipl_count:
        notes_parts.append(
            f"⚠ CONFLICT: {n_banded} banded entries but only {pipl_count} PIPL at this point. "
            f"Max possible banded = {pipl_count}. Please correct."
        )
        status = "conflict"
    elif notes_parts:
        status = "review"
    elif n_banded == 1 and "//" not in entries[0] and "." not in entries[0] and "," not in entries[0]:
        # Single entry with unusual notation — flag for a look
        notes_parts.append("Single entry with non-standard notation — please confirm")
        status = "review"
    else:
        status = "ok"

    notes = "  |  ".join(notes_parts) if notes_parts else ""
    return entries, status, notes


# ── Read 2018 clean file ───────────────────────────────────────────────────────
def read_2018():
    path = CLEAN / "Winter Birds 2018 Clean.xlsx"
    wb   = openpyxl.load_workbook(path)
    ws   = wb["Sheet1"]
    hdrs = [cell.value for cell in ws[1]]

    route_col = next(i for i,h in enumerate(hdrs) if h and "transect" in str(h).lower())
    date_col  = next(i for i,h in enumerate(hdrs) if h and "date" in str(h).lower())

    # Map point columns
    point_map = {}
    for i, h in enumerate(hdrs):
        if not h: continue
        m = re.fullmatch(r"(\d+)Lat",          str(h), re.I); field = "lat"
        if not m: m = re.fullmatch(r"(\d+)Long", str(h), re.I); field = "long"
        if not m: m = re.fullmatch(r"(\d+)PIPL", str(h), re.I); field = "pipl"
        if not m: m = re.fullmatch(r"(\d+)PIPL.?ands", str(h), re.I); field = "band"
        if m:
            n = int(m.group(1))
            point_map.setdefault(n, {})[field] = i

    records = []
    for row in ws.iter_rows(min_row=4, values_only=True):  # rows 1-3 are header/blank
        route = str(row[route_col] or "").strip()
        if not route:
            continue
        date_val = row[date_col]
        try:
            date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val)[:10]
        except Exception:
            date_str = str(date_val)

        for n in sorted(point_map):
            pc = point_map[n]
            pipl_raw = row[pc["pipl"]] if "pipl" in pc else None
            try:
                pipl = int(float(pipl_raw)) if pipl_raw is not None else 0
            except (TypeError, ValueError):
                pipl = 0

            band_raw = row[pc["band"]] if "band" in pc else None
            if band_raw is None or str(band_raw).strip() == "":
                continue
            if is_no_band(str(band_raw).strip()):
                continue

            entries, status, notes = interpret(band_raw, pipl)
            if not entries and status == "ok":
                continue     # genuinely no banded birds — skip row

            records.append({
                "route":       route,
                "point":       n,
                "date":        date_str,
                "pipl":        pipl,
                "raw":         str(band_raw).strip(),
                "entries":     entries,
                "status":      status,
                "notes":       notes,
            })

    return records


# ── Read 2019 clean file ───────────────────────────────────────────────────────
def read_2019():
    path = CLEAN / "Winter Birds 2019 Clean.xlsx"
    wb   = openpyxl.load_workbook(path)
    ws   = wb["Form Responses 1"]
    hdrs = [cell.value for cell in ws[1]]

    route_col = 0
    date_col  = next(i for i,h in enumerate(hdrs) if h and "date" in str(h).lower())

    point_map = {}
    for i, h in enumerate(hdrs):
        if not h: continue
        m = re.search(r'[Pp]oint\s+(\d+)\s+Lat', str(h))
        if m: point_map.setdefault(int(m.group(1)), {})["lat"] = i; continue
        m = re.match(r'^(\d+)\s+Number of Piping', str(h), re.I)
        if m: point_map.setdefault(int(m.group(1)), {})["pipl"] = i; continue
        m = re.match(r'^(\d+)\s+Band/Flag Codes for Piping', str(h), re.I)
        if m: point_map.setdefault(int(m.group(1)), {})["band"] = i; continue

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        route_raw = str(row[route_col] or "").strip()
        if not route_raw: continue
        route = re.sub(r',\s*County=.*$', '', route_raw, flags=re.I).strip()
        date_val = row[date_col]
        try:
            date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val)[:10]
        except Exception:
            date_str = str(date_val)

        for n in sorted(point_map):
            pc = point_map[n]
            pipl_raw = row[pc["pipl"]] if "pipl" in pc else None
            try:
                pipl = int(float(pipl_raw)) if pipl_raw is not None else 0
            except (TypeError, ValueError):
                pipl = 0
            band_raw = row[pc["band"]] if "band" in pc else None
            if not band_raw or str(band_raw).strip() == "" or is_no_band(str(band_raw).strip()):
                continue
            entries, status, notes = interpret(band_raw, pipl)
            if not entries and status == "ok":
                continue
            records.append({"route": route, "point": n, "date": date_str,
                             "pipl": pipl, "raw": str(band_raw).strip(),
                             "entries": entries, "status": status, "notes": notes})
    return records


# ── Read 2020-2024 clean files (Focal Observations + All Species) ──────────────
def read_focal_obs(year: int):
    fname = f"Winter Birds {year} Clean.xlsx"
    wb    = openpyxl.load_workbook(CLEAN / fname)

    # Date/metadata from All Species
    ws_all  = wb["All Species"]
    hdrs_all = [cell.value for cell in ws_all[1]]
    route_col_all = next(i for i,h in enumerate(hdrs_all) if h and 'transect' in str(h).lower())
    date_col_all  = next((i for i,h in enumerate(hdrs_all) if h and 'date' in str(h).lower()), None)

    date_lookup = {}   # normalized_route → date_str
    for row in ws_all.iter_rows(min_row=2, values_only=True):
        r = str(row[route_col_all] or "").strip().rstrip()
        if not r: continue
        dv = row[date_col_all] if date_col_all is not None else None
        try:
            ds = dv.strftime("%Y-%m-%d") if hasattr(dv, "strftime") else str(dv)[:10]
        except Exception:
            ds = str(dv) if dv else ""
        date_lookup[r.lower()] = ds

    # Band + PIPL from Focal Observations
    ws_focal = wb["Focal Observations"]
    hdrs_foc = [cell.value for cell in ws_focal[1]]
    route_col_foc = next(i for i,h in enumerate(hdrs_foc) if h and 'transect' in str(h).lower())

    point_map = {}
    for i, h in enumerate(hdrs_foc):
        if not h: continue
        m = re.search(r'Number of Piping Plovers \(point\s+(\d+)\)', str(h), re.I)
        if m: point_map.setdefault(int(m.group(1)), {})["pipl"] = i; continue
        m = re.search(r'PIPL Band/Flag Codes \(point\s+(\d+)\)', str(h), re.I)
        if m: point_map.setdefault(int(m.group(1)), {})["band"] = i; continue

    records = []
    for row in ws_focal.iter_rows(min_row=2, values_only=True):
        route = str(row[route_col_foc] or "").strip().rstrip()
        if not route: continue
        date_str = date_lookup.get(route.lower(), "")

        for n in sorted(point_map):
            pc = point_map[n]
            pipl_raw = row[pc["pipl"]] if "pipl" in pc else None
            try:
                pipl = int(float(pipl_raw)) if pipl_raw is not None else 0
            except (TypeError, ValueError):
                pipl = 0
            band_raw = row[pc["band"]] if "band" in pc else None
            if not band_raw or str(band_raw).strip() == "" or is_no_band(str(band_raw).strip()):
                continue
            entries, status, notes = interpret(band_raw, pipl)
            if not entries and status == "ok":
                continue
            records.append({"route": route, "point": n, "date": date_str,
                             "pipl": pipl, "raw": str(band_raw).strip(),
                             "entries": entries, "status": status, "notes": notes})
    return records


# ── Build the Excel workbook ───────────────────────────────────────────────────
def make_workbook(records_by_year: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Sheet 1: Instructions ─────────────────────────────────────────────────
    ws_inst = wb.create_sheet("How to Fill This")
    ws_inst.sheet_view.showGridLines = False

    inst_header_fill = PatternFill("solid", fgColor=C_INST_HDR)
    inst_body_fill   = PatternFill("solid", fgColor=C_INST_BG)
    white_fill       = PatternFill("solid", fgColor="FFFFFF")
    bold_white       = Font(bold=True, color="FFFFFF", size=13)
    bold_dark        = Font(bold=True, color="1B4F72", size=11)
    normal           = Font(size=11)
    wrap             = Alignment(wrap_text=True, vertical="top")

    def inst_row(ws, row_num, text, is_header=False, is_blank=False):
        cell = ws.cell(row=row_num, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if is_header:
            cell.fill = inst_header_fill
            cell.font = bold_white
        elif is_blank:
            cell.fill = white_fill
        else:
            cell.fill = inst_body_fill
            cell.font = normal
        ws.row_dimensions[row_num].height = 30 if is_header else 22

    ws_inst.column_dimensions["A"].width = 3
    ws_inst.column_dimensions["B"].width = 90

    content = [
        ("USFWS Piping Plover Winter Survey — Band Data Review (2018–2024)", True),
        ("", False),
        ("Thank you for reviewing this file. We have processed the raw banding notes from the "
         "2018–2024 winter surveys and split each cell into individual bird entries. "
         "Each year is a separate sheet. Your expertise is needed to confirm or correct our work.", False),
        ("", False),
        ("HOW TO FILL THIS IN — please read carefully", True),
        ("", False),
        ("Each row is one banding cell from the original data sheet "
         "(one route + one point + one survey).", False),
        ("", False),
        ("Column E  'Original Text'     — Exactly what was written in the field data. Read-only.", False),
        ("Column F  'Our Interpretation'— How we split it into individual birds. Read-only.", False),
        ("Column G  'Our Notes'         — Any flag or issue we found. Read-only.", False),
        ("", False),
        ("Column H  'Your Correction'   — THE ONLY COLUMN YOU NEED TO FILL IN.", False),
        ("", False),
        ("  ✓  If Column F looks correct  →  leave Column H BLANK. No action needed.", False),
        ("", False),
        ("  ✗  If Column F is wrong  →  write the correct version in Column H\n"
         "     using numbered bullets, one bird per line, like this:\n"
         "\n"
         "         1) Of//GG:S//K\n"
         "         2) Gf//OO:X//GB\n"
         "         3) Yf(O64)//WK:S//bW\n"
         "\n"
         "     Each line = one individual bird's band combination.\n"
         "     If there is only 1 bird, just write:  1) <band combo>", False),
        ("", False),
        ("COLOUR GUIDE", True),
        ("", False),
        ("  White rows  — We are confident. A quick glance is enough.", False),
        ("  Yellow rows — We are unsure about the count or split. Please check carefully.", False),
        ("  Orange rows — Count conflict (more banded entries than total PIPL at that point).\n"
         "                These definitely need your correction.", False),
        ("", False),
        ("ONE MORE THING", True),
        ("", False),
        ("Column D 'PIPL at Point' is the total Piping Plovers counted at that location.\n"
         "The number of banded birds in Column H cannot exceed this number.\n"
         "For example: if PIPL at Point = 3, you can have at most 3 banded entries.", False),
        ("", False),
        ("Thank you — please email the completed file back when done.", False),
    ]

    for i, (text, is_hdr) in enumerate(content, start=2):
        inst_row(ws_inst, i, text, is_header=is_hdr, is_blank=(text == "" and not is_hdr))

    # ── Sheet 2+: one per year ─────────────────────────────────────────────────
    STATUS_FILL = {
        "ok":       PatternFill("solid", fgColor=C_OK),
        "review":   PatternFill("solid", fgColor=C_REVIEW),
        "conflict": PatternFill("solid", fgColor=C_CONFLICT),
    }

    HDR_COLS = ["Route", "Point", "Date", "PIPL at Point",
                "Original Text", "Our Interpretation",
                "Our Notes (read only)", "Your Correction — leave blank if Col F is correct"]
    COL_WIDTHS = [38, 7, 12, 14, 45, 45, 38, 45]

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor=C_HEADER)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(border_style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for year, records in records_by_year.items():
        ws = wb.create_sheet(str(year))
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        # Header row
        for ci, (col_name, width) in enumerate(zip(HDR_COLS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = cell_border
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 28

        # Data rows
        for ri, rec in enumerate(records, start=2):
            interpretation = "\n".join(
                f"{i+1}) {e}" for i, e in enumerate(rec["entries"])
            ) if rec["entries"] else "(no banded birds — all unbanded)"

            values = [
                rec["route"],
                rec["point"],
                rec["date"],
                rec["pipl"],
                rec["raw"],
                interpretation,
                rec["notes"],
                "",             # Correction — biologist fills this
            ]
            fill = STATUS_FILL[rec["status"]]
            for ci, val in enumerate(values, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill      = fill
                cell.border    = cell_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Auto row height (rough estimate)
            n_lines = max(
                len(str(rec["raw"]).split("\n")),
                len(rec["entries"]) + 1,
                1,
            )
            ws.row_dimensions[ri].height = max(18, n_lines * 16)

    return wb


# ── Run ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    readers = {
        2018: read_2018,
        2019: read_2019,
        2020: lambda: read_focal_obs(2020),
        2021: lambda: read_focal_obs(2021),
        2022: lambda: read_focal_obs(2022),
        2023: lambda: read_focal_obs(2023),
        2024: lambda: read_focal_obs(2024),
    }

    records_by_year = {}
    total_ok = total_review = total_conflict = 0

    for year, reader in readers.items():
        print(f"Reading {year}...")
        recs = reader()
        records_by_year[str(year)] = recs
        ok       = sum(1 for r in recs if r["status"] == "ok")
        review   = sum(1 for r in recs if r["status"] == "review")
        conflict = sum(1 for r in recs if r["status"] == "conflict")
        total_ok += ok; total_review += review; total_conflict += conflict
        print(f"  {len(recs):3d} cells  —  "
              f"White: {ok}  Yellow: {review}  Orange: {conflict}")

    print(f"\nAll years total — White: {total_ok}  "
          f"Yellow: {total_review}  Orange: {total_conflict}")

    print("\nBuilding Excel workbook...")
    wb = make_workbook(records_by_year)
    wb.save(OUT)
    print(f"Saved: {OUT}")
