"""
Step 6: Final Report Generation
─────────────────────────────────
Combines all pipeline outputs into a single professional Excel file
with 3 sheets:

  Sheet 1 — Clean_Data
      The fully processed, validated rows ready for biologist use.

  Sheet 2 — Removed_Rows
      Every row that was removed at any step, with columns:
        unique_id, source_file, source_sheet, removal_reason,
        + all original data columns

  Sheet 3 — Summary_Report
      Processing statistics with separate sections for each removal
      category (geography, missing fields, duplicates), coordinate
      corrections, location warnings, and coastal buffer warnings.

Output: database4_FINAL.xlsx
"""

import pandas as pd
import os
import sys
from datetime import datetime

# ── Load config ────────────────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from database4_config import config

# ── Resolve paths ──────────────────────────────────────────────────────────────
script_dir    = os.path.dirname(os.path.abspath(__file__))
output_folder = os.path.normpath(os.path.join(script_dir, config["output"]["folder"]))

clean_path   = os.path.join(output_folder, "database4_clean.xlsx")
removed_path = os.path.join(output_folder, "database4_removed.xlsx")
ids_path     = os.path.join(output_folder, "database4_with_ids.xlsx")
final_path   = os.path.join(output_folder, "database4_FINAL.xlsx")

# ── Load data ──────────────────────────────────────────────────────────────────
for path, label in [(clean_path, "Clean data (Step 5)"), (removed_path, "Removed rows"), (ids_path, "Original with IDs (Step 1)")]:
    if not os.path.exists(path):
        print(f"[ERROR] Missing file: {path} ({label})")
        print("        Run all previous steps first.")
        sys.exit(1)

clean_df   = pd.read_excel(clean_path)
removed_df = pd.read_excel(removed_path)
original   = pd.read_excel(ids_path)

print(f"  Clean rows   : {len(clean_df)}")
print(f"  Removed rows : {len(removed_df)}")
print(f"  Total input  : {len(original)}")

# ── Extract internal columns before dropping them from Clean_Data ─────────────
geo_warnings = []
if "_geo_warning" in clean_df.columns:
    warned = clean_df[clean_df["_geo_warning"].notna()]
    for _, row in warned.iterrows():
        geo_warnings.append({
            "unique_id": row.get("unique_id"),
            "warning":   row["_geo_warning"],
        })
    clean_df = clean_df.drop(columns=["_geo_warning"])

geo_corrections = []
if "_geo_correction" in clean_df.columns:
    corrected = clean_df[clean_df["_geo_correction"].notna()]
    for _, row in corrected.iterrows():
        geo_corrections.append({
            "unique_id":  row.get("unique_id"),
            "correction": row["_geo_correction"],
        })
    clean_df = clean_df.drop(columns=["_geo_correction"])

# ── Sheet 2: Removed_Rows ──────────────────────────────────────────────────────
removed_sheet = removed_df.copy()
removed_sheet = removed_sheet.rename(columns={"_removal_reason": "removal_reason"})

# Drop internal columns from removed sheet too
for col in ["_geo_warning", "_geo_correction"]:
    if col in removed_sheet.columns:
        removed_sheet = removed_sheet.drop(columns=[col])

front_cols = ["unique_id", "removal_reason", "source_database", "source_file", "source_sheet"]
clean_cols = [c for c in clean_df.columns
              if c not in front_cols and c in removed_sheet.columns]

final_removed_cols = (
    [c for c in front_cols if c in removed_sheet.columns] +
    clean_cols
)
removed_sheet = removed_sheet[final_removed_cols]

# ── Sheet 3: Summary_Report ────────────────────────────────────────────────────
total_input   = len(original)
total_clean   = len(clean_df)
total_removed = len(removed_df)

# ── Categorize removal reasons ────────────────────────────────────────────────
geo_removals     = removed_df[removed_df["_removal_reason"].str.startswith("Outside Florida", na=False)]
missing_removals = removed_df[removed_df["_removal_reason"].str.startswith("Missing", na=False)]
dup_removals     = removed_df[removed_df["_removal_reason"].str.startswith("Duplicate", na=False)]

geo_reason_counts = geo_removals["_removal_reason"].value_counts().reset_index()
geo_reason_counts.columns = ["removal_reason", "count"]

missing_reason_counts = missing_removals["_removal_reason"].value_counts().reset_index()
missing_reason_counts.columns = ["removal_reason", "count"]

# Partial location warnings (rows kept but missing some location fields)
loc_fields = config["location_fields"]["fields"]
partial_warnings = []
for _, row in clean_df.iterrows():
    missing = [f for f in loc_fields if f in clean_df.columns and (pd.isna(row.get(f)) or str(row.get(f)).strip() == "")]
    if 0 < len(missing) < len(loc_fields):
        partial_warnings.append({
            "unique_id":      row.get("unique_id"),
            "missing_fields": ", ".join(missing)
        })

# ── Build summary rows ───────────────────────────────────────────────────────
SECTION = "§"

summary_rows = [
    ("Run date",                                            datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("Database",                                            config["database_name"]),
    ("Source database label",                               config["source_database"]),
    (f"{SECTION}Input",                                     ""),
    ("Total rows input",                                    total_input),
    ("Source files",                                        ", ".join(config["input"]["files"])),
    (f"{SECTION}Output",                                    ""),
    ("Total rows in Clean_Data",                            total_clean),
    ("Total rows removed",                                  total_removed),
    ("Percentage of rows kept",                             f"{round(total_clean / total_input * 100, 2)}%"),
]

# ── Geography Removals ────────────────────────────────────────────────────────
summary_rows.append((f"{SECTION}Geography Removals", ""))
summary_rows.append(("Total removed for geography", len(geo_removals)))

if len(geo_removals) > 0:
    for _, r in geo_reason_counts.iterrows():
        summary_rows.append((f"  {r['removal_reason']}", f"{r['count']} rows"))
else:
    summary_rows.append(("  None", ""))

# ── Missing Field Removals ────────────────────────────────────────────────────
summary_rows.append((f"{SECTION}Missing Field Removals", ""))
summary_rows.append(("Total removed for missing fields", len(missing_removals)))

if len(missing_removals) > 0:
    for _, r in missing_reason_counts.iterrows():
        summary_rows.append((f"  {r['removal_reason']}", f"{r['count']} rows"))
else:
    summary_rows.append(("  None", ""))

# ── Duplicate Removals ────────────────────────────────────────────────────────
summary_rows.append((f"{SECTION}Duplicate Removals", ""))
summary_rows.append(("Total duplicates removed", len(dup_removals)))
summary_rows.append(("Criteria used", ", ".join(config["duplicate_criteria"])))

# ── Coordinate Corrections ────────────────────────────────────────────────────
summary_rows.append((f"{SECTION}Coordinate Corrections", ""))
summary_rows.append(("Rows with auto-corrected coordinates", len(geo_corrections)))

if geo_corrections:
    for c in geo_corrections:
        summary_rows.append((f"  unique_id {c['unique_id']}", c["correction"]))
else:
    summary_rows.append(("  None", ""))

# ── Location Warnings ─────────────────────────────────────────────────────────
summary_rows.append((f"{SECTION}Location Warnings", ""))
summary_rows.append(("Rows with partial location data (kept but flagged)", len(partial_warnings)))

if partial_warnings:
    for w in partial_warnings:
        summary_rows.append((f"  unique_id {w['unique_id']}", f"Missing: {w['missing_fields']}"))
else:
    summary_rows.append(("  None", ""))

# ── Coastal Buffer Warnings ───────────────────────────────────────────────────
summary_rows.append((f"{SECTION}Coastal Buffer Warnings", ""))
summary_rows.append(("Rows kept but within coastal buffer (not on land)", len(geo_warnings)))

if geo_warnings:
    for w in geo_warnings:
        summary_rows.append((f"  unique_id {w['unique_id']}", w["warning"]))
else:
    summary_rows.append(("  None", ""))

summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

# ── Write final Excel with 3 sheets ───────────────────────────────────────────
with pd.ExcelWriter(final_path, engine="openpyxl") as writer:
    clean_df.to_excel(writer,    sheet_name="Clean_Data",     index=False)
    removed_sheet.to_excel(writer, sheet_name="Removed_Rows", index=False)
    summary_df.to_excel(writer,  sheet_name="Summary_Report", index=False)

    # ── Formatting ─────────────────────────────────────────────────────────────
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    wb = writer.book

    header_font  = Font(bold=True, color="FFFFFF")
    section_font = Font(bold=True, color="1565C0")
    header_fills = {
        "Clean_Data":     PatternFill("solid", fgColor="2E7D32"),  # dark green
        "Removed_Rows":   PatternFill("solid", fgColor="C62828"),  # dark red
        "Summary_Report": PatternFill("solid", fgColor="1565C0"),  # dark blue
    }

    for sheet_name in ["Clean_Data", "Removed_Rows", "Summary_Report"]:
        ws   = wb[sheet_name]
        fill = header_fills[sheet_name]

        # Style header row
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")

        # Apply Excel date format to date columns
        if sheet_name in ["Clean_Data", "Removed_Rows"]:
            for col in ws.iter_cols(min_row=1, max_row=1):
                if col[0].value == "ResightDate":
                    col_letter = get_column_letter(col[0].column)
                    for row in ws[col_letter]:
                        if row.row > 1:
                            row.number_format = "MM/DD/YYYY"

        # Summary_Report: bold section header rows (marked with §)
        if sheet_name == "Summary_Report":
            for row in ws.iter_rows(min_row=2):
                cell = row[0]
                if cell.value and str(cell.value).startswith(SECTION):
                    cell.value = str(cell.value).replace(SECTION, "")
                    for c in row:
                        c.font = section_font
                        c.fill = PatternFill("solid", fgColor="E3F2FD")  # light blue bg

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

print(f"\n[DONE] Step 6 complete")
print(f"  Final report : {final_path}")
print(f"  Sheets       : Clean_Data ({total_clean} rows) | Removed_Rows ({total_removed} rows) | Summary_Report")
