"""
Step 6: Final Report Generation
─────────────────────────────────
Combines all pipeline outputs into a single professional Excel file
with 3 sheets:

  Sheet 1 — Clean_Data        : fully processed, validated rows
  Sheet 2 — Removed_Rows      : every removed row with reason + original data
  Sheet 3 — Summary_Report    : processing statistics and warnings

Output: database1_FINAL.xlsx
"""

import pandas as pd
import os
import sys
from datetime import datetime

# ── Load config ────────────────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from database1_config import config

# ── Resolve paths ──────────────────────────────────────────────────────────────
script_dir    = os.path.dirname(os.path.abspath(__file__))
output_folder = os.path.normpath(os.path.join(script_dir, config["output"]["folder"]))

clean_path   = os.path.join(output_folder, "database1_clean.xlsx")
removed_path = os.path.join(output_folder, "database1_removed.xlsx")
ids_path     = os.path.join(output_folder, "database1_with_ids.xlsx")
final_path   = os.path.join(output_folder, "database1_FINAL.xlsx")

# ── Load data ──────────────────────────────────────────────────────────────────
for path, label in [
    (clean_path,   "Clean data (Step 5)"),
    (removed_path, "Removed rows"),
    (ids_path,     "Original with IDs (Step 1)"),
]:
    if not os.path.exists(path):
        print(f"[ERROR] Missing file: {path} ({label})")
        print("        Run all previous steps first.")
        sys.exit(1)

clean_df   = pd.read_excel(clean_path)
removed_df = pd.read_excel(removed_path)
original   = pd.read_excel(ids_path)

print(f"  Clean rows   : {len(clean_df)}")
print(f"  Removed rows : {len(removed_df)}")
print(f"  Total input  : {len(original)}")

# ── Sheet 2: Removed_Rows ──────────────────────────────────────────────────────
removed_sheet = removed_df.copy()
removed_sheet = removed_sheet.rename(columns={"_removal_reason": "removal_reason"})

front_cols = ["unique_id", "removal_reason", "source_database", "source_file", "source_sheet"]
clean_cols = [c for c in clean_df.columns if c not in front_cols and c in removed_sheet.columns]

final_removed_cols = (
    [c for c in front_cols if c in removed_sheet.columns] +
    clean_cols
)
removed_sheet = removed_sheet[final_removed_cols]

# ── Sheet 3: Summary_Report ────────────────────────────────────────────────────
total_input   = len(original)
total_clean   = len(clean_df)
total_removed = len(removed_df)

reason_counts = removed_df["_removal_reason"].value_counts().reset_index()
reason_counts.columns = ["removal_reason", "count"]
reason_counts["percentage"] = (reason_counts["count"] / total_input * 100).round(2).astype(str) + "%"

# Partial location warnings
loc_fields = config["location_fields"]["fields"]
renames    = config.get("column_rename", {})
partial_warnings = []
for _, row in clean_df.iterrows():
    missing = [
        f for f in loc_fields
        if renames.get(f, f) in clean_df.columns
        and (pd.isna(row.get(renames.get(f, f))) or str(row.get(renames.get(f, f))).strip() == "")
    ]
    if 0 < len(missing) < len(loc_fields):
        partial_warnings.append({
            "unique_id":      row.get("unique_id"),
            "missing_fields": ", ".join(missing),
        })

SECTION = "§"

summary_rows = [
    ("Run date",                                            datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("Database",                                            config["database_name"]),
    ("Source database label",                               config["source_database"]),
    (f"{SECTION}Input",                                     ""),
    ("Total rows input",                                    total_input),
    ("Source files",                                        ", ".join(config["input"]["files"])),
    (f"{SECTION}Output",                                    ""),
    ("Total rows in Clean_Data",                            total_clean),
    ("Total rows removed",                                  total_removed),
    ("Percentage of rows kept",                             f"{round(total_clean / total_input * 100, 2)}%"),
    (f"{SECTION}Removal Breakdown",                         ""),
]

for _, r in reason_counts.iterrows():
    summary_rows.append((f"  {r['removal_reason']}", f"{r['count']} rows ({r['percentage']})"))

summary_rows.append((f"{SECTION}Location Warnings", ""))
summary_rows.append(("Rows with partial location data (kept but flagged)", len(partial_warnings)))

if partial_warnings:
    for w in partial_warnings:
        summary_rows.append((f"  unique_id {w['unique_id']}", f"Missing: {w['missing_fields']}"))
else:
    summary_rows.append(("  None", ""))

summary_rows.append((f"{SECTION}Duplicate Criteria", ""))
summary_rows.append(("Columns used to detect duplicates", ", ".join(config["duplicate_criteria"])))
summary_rows.append(("Note", "Comment fields excluded — rows with different comments can still be duplicates"))

summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

# ── Write final Excel with 3 sheets ───────────────────────────────────────────
with pd.ExcelWriter(final_path, engine="openpyxl") as writer:
    clean_df.to_excel(writer,      sheet_name="Clean_Data",     index=False)
    removed_sheet.to_excel(writer, sheet_name="Removed_Rows",   index=False)
    summary_df.to_excel(writer,    sheet_name="Summary_Report", index=False)

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = writer.book

    header_font  = Font(bold=True, color="FFFFFF")
    section_font = Font(bold=True, color="1565C0")
    header_fills = {
        "Clean_Data":     PatternFill("solid", fgColor="2E7D32"),  # dark green
        "Removed_Rows":   PatternFill("solid", fgColor="C62828"),  # dark red
        "Summary_Report": PatternFill("solid", fgColor="1565C0"),  # dark blue
    }

    # Date columns to format (after rename)
    date_cols = {"ObservationDate"}

    for sheet_name in ["Clean_Data", "Removed_Rows", "Summary_Report"]:
        ws   = wb[sheet_name]
        fill = header_fills[sheet_name]

        # Style header row
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")

        # Apply Excel date format to date columns
        if sheet_name in ["Clean_Data", "Removed_Rows"]:
            for col in ws.iter_cols(min_row=1, max_row=1):
                if col[0].value in date_cols:
                    col_letter = get_column_letter(col[0].column)
                    for row in ws[col_letter]:
                        if row.row > 1:
                            row.number_format = "MM/DD/YYYY"

        # Summary_Report: bold section header rows
        if sheet_name == "Summary_Report":
            for row in ws.iter_rows(min_row=2):
                cell = row[0]
                if cell.value and str(cell.value).startswith(SECTION):
                    cell.value = str(cell.value).replace(SECTION, "")
                    for c in row:
                        c.font = section_font
                        c.fill = PatternFill("solid", fgColor="E3F2FD")

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

print(f"\n[DONE] Step 6 complete")
print(f"  Final report : {final_path}")
print(f"  Sheets       : Clean_Data ({total_clean} rows) | Removed_Rows ({total_removed} rows) | Summary_Report")
