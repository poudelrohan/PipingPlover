"""
combine_all_databases.py
─────────────────────────
Combines the FINAL files of Database 1, 2, 3, and 4 into one workbook for
biologist review.

Inputs (one per database):
  Databases/Database1/Output/database1_FINAL.xlsx          → Clean_Data sheet
  Databases/Database2/Output/database2_FINAL.xlsx          → Clean_Data sheet
  Databases/Database3/Output/AllYears/db3_ALL_YEARS_FINAL.xlsx → All_Years_Combined sheet
  Databases/Database4/Output/database4_FINAL.xlsx          → Clean_Data sheet

Output:
  Databases/AllDatabasesCombined/db_ALL_COMBINED_FINAL.xlsx
  ├─ AllDBCombined  — harmonized 35-column view of all four databases stacked
  └─ DB1, DB2, DB3, DB4 — exact copies of each source database's main data sheet
     (no Removed_Rows, no Summary sheets carried over)

Harmonization rules for the AllDBCombined sheet:
  • Shared concepts (Date, Location, Latitude, Longitude, TotalObserved, Observer,
    Comments) are mapped to unified column names regardless of source DB.
  • DB-specific columns get a "(DB1)/(DB2)/(DB3)/(DB4)" suffix in their name so
    biologists immediately see which database produced them.
  • Species: all rows should be PIPL. DB1's "Piping Plover" is normalised to
    "PIPL". Any non-PIPL species is flagged in chat.
  • A new global unique_id (1..N) is assigned. The original per-DB unique_id
    is preserved inside db_id (e.g. db1_00042, db3_2013_0001, db4_017).
"""

import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# Paths
# ══════════════════════════════════════════════════════════════════════════════

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT  = os.path.normpath(os.path.join(SCRIPT_DIR, ".."))

DB_SOURCES = {
    1: {
        "path":  os.path.join(REPO_ROOT, "Databases/Database1/Output/database1_FINAL.xlsx"),
        "sheet": "Clean_Data",
    },
    2: {
        "path":  os.path.join(REPO_ROOT, "Databases/Database2/Output/database2_FINAL.xlsx"),
        "sheet": "Clean_Data",
    },
    3: {
        "path":  os.path.join(REPO_ROOT, "Databases/Database3/Output/AllYears/db3_ALL_YEARS_FINAL.xlsx"),
        "sheet": "All_Years_Combined",
    },
    4: {
        "path":  os.path.join(REPO_ROOT, "Databases/Database4/Output/database4_FINAL.xlsx"),
        "sheet": "Clean_Data",
    },
}

OUT_DIR  = os.path.join(REPO_ROOT, "Databases/AllDatabasesCombined")
OUT_PATH = os.path.join(OUT_DIR, "db_ALL_COMBINED_FINAL.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# Final column order for AllDBCombined sheet (BASE names — no DB-suffix yet)
# Suffixes are computed dynamically after concatenation based on which DBs
# actually populate each column.
# ══════════════════════════════════════════════════════════════════════════════

FINAL_COLS = [
    # Identifiers (added by this script — always populated everywhere)
    "unique_id", "database", "db_id", "Year",
    # Core shared fields
    "Date", "StartTime", "EndTime",
    "Location", "Latitude", "Longitude", "GroupNumber",
    "Species", "TotalObserved", "TotalBanded",
    "Observer", "ObserverEmail",
    # Band/flag info
    "FlagCode", "FlagColor", "BandCombo",
    "FlagID", "UpperLeft", "LowerLeft", "UpperRight", "LowerRight",
    # Situational
    "HabitatType", "Tide", "Foraging", "Roosting",
    "FlockActivity", "WeatherCondition",
    # Comments + source
    "PrimaryComments", "SecondaryComments",
    "source_database", "source_file", "source_sheet",
]

# Identifier columns are added by this script and should never get a DB suffix
NEVER_SUFFIX = {"unique_id", "database", "db_id", "Year"}


# ══════════════════════════════════════════════════════════════════════════════
# Per-DB harmonization
# ══════════════════════════════════════════════════════════════════════════════

def _empty_row():
    return {c: pd.NA for c in FINAL_COLS}


def _safe_year_from_date(d):
    if d is None or (isinstance(d, float) and pd.isna(d)):
        return pd.NA
    try:
        return pd.to_datetime(d).year
    except Exception:
        return pd.NA


def _normalize_species(value, dbnum: int, warnings: list) -> str:
    """Convert various species spellings to the 4-letter code. Warn on non-PIPL."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return pd.NA
    s = str(value).strip()
    if s.upper() == "PIPL":
        return "PIPL"
    if s.lower() == "piping plover":
        return "PIPL"
    # Anything else is a warning
    warnings.append(f"DB{dbnum}: non-PIPL species value {value!r} encountered")
    return s.upper()


def compute_dynamic_suffixes(df_combined: pd.DataFrame) -> dict:
    """
    For each base column, find which DBs (1/2/3/4) populate it with at least
    one non-null value, then build the new column name:
        - If all 4 DBs contribute → keep the base name as-is
        - If a subset contributes → append " (DBx, DBy)"
        - NEVER_SUFFIX columns always keep the base name
    Returns {old_name → new_name} mapping.
    """
    rename_map = {}
    for col in df_combined.columns:
        if col in NEVER_SUFFIX:
            rename_map[col] = col
            continue
        contribs = []
        for db_num in (1, 2, 3, 4):
            sub = df_combined[df_combined["database"] == db_num]
            if len(sub) and sub[col].notna().any():
                contribs.append(db_num)
        if not contribs:
            # No DB populates it — leave as-is (shouldn't happen for our schema)
            rename_map[col] = col
        elif len(contribs) == 4:
            rename_map[col] = col   # everyone contributes → no suffix
        else:
            suffix = ", ".join(f"DB{n}" for n in contribs)
            rename_map[col] = f"{col} ({suffix})"
    return rename_map


def harmonize_db1(df: pd.DataFrame, warnings: list) -> pd.DataFrame:
    """eBird observations → 35-col harmonized rows."""
    out = []
    for _, r in df.iterrows():
        row = _empty_row()
        row["database"]       = 1
        row["db_id"]          = f"db1_{int(r['unique_id']):05d}"
        row["Date"]           = r.get("ObservationDate")
        row["Year"]           = _safe_year_from_date(r.get("ObservationDate"))
        row["StartTime"]      = r.get("TimeStarted")
        row["Location"]       = r.get("LocationName")
        row["Latitude"]       = r.get("Latitude")
        row["Longitude"]      = r.get("Longitude")
        row["Species"]        = _normalize_species(r.get("CommonName"), 1, warnings)
        # ObservationCount is sometimes a string in DB1 — coerce safely
        try:
            row["TotalObserved"] = int(float(r.get("ObservationCount")))
        except (TypeError, ValueError):
            row["TotalObserved"] = pd.NA
        row["PrimaryComments"]   = r.get("ChecklistComments")
        row["SecondaryComments"] = r.get("SpeciesComments")
        row["source_database"]   = r.get("source_database")
        row["source_file"]       = r.get("source_file")
        row["source_sheet"]      = r.get("source_sheet")
        out.append(row)
    return pd.DataFrame(out, columns=FINAL_COLS)


def harmonize_db2(df: pd.DataFrame, warnings: list) -> pd.DataFrame:
    """Non-breeding PIPL survey → 35-col harmonized rows."""
    out = []
    for _, r in df.iterrows():
        row = _empty_row()
        row["database"]            = 2
        row["db_id"]               = f"db2_{int(r['unique_id']):04d}"
        row["Date"]                = r.get("SurveyDate")
        row["Year"]                = _safe_year_from_date(r.get("SurveyDate"))
        row["StartTime"]           = r.get("TimeSited")
        row["Location"]            = r.get("Route")
        row["Latitude"]            = r.get("Latitude")
        row["Longitude"]           = r.get("Longitude")
        row["GroupNumber"]         = r.get("GroupNumber")
        row["Species"]             = "PIPL"
        row["TotalObserved"]       = r.get("TotalObserved")
        row["TotalBanded"]         = r.get("TotalBanded")
        row["Observer"]            = r.get("Observer")
        row["HabitatType"]   = r.get("HabitatType")
        row["Tide"]          = r.get("Tide")
        row["Foraging"]      = r.get("Foraging")
        row["Roosting"]      = r.get("Roosting")
        row["PrimaryComments"]     = r.get("Notes")
        row["source_database"]     = r.get("source_database")
        row["source_file"]         = r.get("source_file")
        row["source_sheet"]        = r.get("source_sheet")
        out.append(row)
    return pd.DataFrame(out, columns=FINAL_COLS)


def harmonize_db3(df: pd.DataFrame, warnings: list) -> pd.DataFrame:
    """Winter Bird Survey (all years) → 35-col harmonized rows."""
    out = []
    for _, r in df.iterrows():
        row = _empty_row()
        row["database"]                  = 3
        # DB3 already has year_id like "2013_0001"; prefix with db3_
        row["db_id"]                     = f"db3_{r.get('year_id')}"
        row["Year"]                      = r.get("SurveyYear")
        row["Date"]                      = r.get("SurveyDate")
        row["StartTime"]                 = r.get("SurveyTime")
        row["Location"]                  = r.get("Route")
        row["Latitude"]                  = r.get("Latitude")
        row["Longitude"]                 = r.get("Longitude")
        row["GroupNumber"]               = r.get("GroupNumber")
        row["Species"]                   = "PIPL"
        row["TotalObserved"]             = r.get("TotalObserved")
        # TotalBanded: 1 if BandCombo non-null (this row is a banded bird); else 0
        row["TotalBanded"]               = 1 if pd.notna(r.get("BandCombo")) else 0
        row["Observer"]                  = r.get("Observer")
        row["ObserverEmail"]       = r.get("ObserverEmail")
        row["FlagCode"]                  = r.get("FlagCode")
        row["FlagColor"]           = r.get("FlagColor")
        row["BandCombo"]           = r.get("BandCombo")
        row["WeatherCondition"]    = r.get("WeatherCondition")
        row["PrimaryComments"]           = r.get("Comments")
        row["source_database"]           = r.get("source_database")
        row["source_file"]               = r.get("source_file")
        row["source_sheet"]              = r.get("source_sheet")
        out.append(row)
    return pd.DataFrame(out, columns=FINAL_COLS)


def harmonize_db4(df: pd.DataFrame, warnings: list) -> pd.DataFrame:
    """Banded bird resights → 35-col harmonized rows."""
    out = []
    for _, r in df.iterrows():
        row = _empty_row()
        row["database"]              = 4
        row["db_id"]                 = f"db4_{int(r['unique_id']):03d}"
        row["Date"]                  = r.get("ResightDate")
        row["Year"]                  = _safe_year_from_date(r.get("ResightDate"))
        row["StartTime"]             = r.get("StartTime")
        row["EndTime"]         = r.get("EndTime")
        row["Location"]              = r.get("LocationName")
        row["Latitude"]              = r.get("Latitude")
        row["Longitude"]             = r.get("Longitude")
        row["Species"]               = _normalize_species(r.get("SpeciesID"), 4, warnings)
        # Each DB4 row IS a banded bird, count from FlockSize if present, else 1
        try:
            row["TotalObserved"] = int(float(r.get("FlockSize")))
        except (TypeError, ValueError):
            row["TotalObserved"] = 1
        row["TotalBanded"]           = 1
        # Combine first + last name into a single Observer field
        first = str(r.get("ObserverFirst") or "").strip()
        last  = str(r.get("ObserverLast") or "").strip()
        row["Observer"] = (f"{first} {last}".strip()) or pd.NA
        row["FlagCode"]              = r.get("FlagCode")
        row["FlagID"]          = r.get("FlagID")
        row["UpperLeft"]       = r.get("UpperLeft")
        row["LowerLeft"]       = r.get("LowerLeft")
        row["UpperRight"]      = r.get("UpperRight")
        row["LowerRight"]      = r.get("LowerRight")
        row["FlockActivity"]   = r.get("FlockActivityID")
        row["PrimaryComments"]       = r.get("MasterComments")
        row["SecondaryComments"]     = r.get("ResightingComments")
        row["source_database"]       = r.get("source_database")
        row["source_file"]           = r.get("source_file")
        row["source_sheet"]          = r.get("source_sheet")
        out.append(row)
    return pd.DataFrame(out, columns=FINAL_COLS)


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    print(f"\n{'='*70}")
    print("Combining all 4 databases into one workbook")
    print(f"  Output: {OUT_PATH}")
    print()

    # ── Load each DB's main data sheet ─────────────────────────────────────────
    raw_dfs = {}
    for n, info in DB_SOURCES.items():
        df = pd.read_excel(info["path"], sheet_name=info["sheet"])
        raw_dfs[n] = df
        print(f"  DB{n}: {len(df):>6} rows loaded from {info['sheet']!r}")

    # ── Harmonize each into the 35-col layout ─────────────────────────────────
    warnings = []
    harm_dfs = {
        1: harmonize_db1(raw_dfs[1], warnings),
        2: harmonize_db2(raw_dfs[2], warnings),
        3: harmonize_db3(raw_dfs[3], warnings),
        4: harmonize_db4(raw_dfs[4], warnings),
    }

    # ── Concatenate in DB order: DB1, DB2, DB3, DB4 ───────────────────────────
    combined = pd.concat(
        [harm_dfs[1], harm_dfs[2], harm_dfs[3], harm_dfs[4]],
        ignore_index=True, sort=False,
    )
    combined["unique_id"] = range(1, len(combined) + 1)
    combined = combined[FINAL_COLS]   # enforce column order

    # Strip time component from Date column — biologists want date-only display.
    # pd.to_datetime → .dt.normalize keeps it datetime-typed but zeros the time;
    # then we set an Excel cell format below so it renders as YYYY-MM-DD.
    combined["Date"] = pd.to_datetime(combined["Date"], errors="coerce").dt.normalize()

    # Capture stats BEFORE rename so we can still reference base column names
    total_pipl   = combined["TotalObserved"].fillna(0).astype(float).sum()
    total_banded = combined["TotalBanded"].fillna(0).astype(float).sum()

    # ── Apply dynamic DB-suffix to column names ───────────────────────────────
    rename_map = compute_dynamic_suffixes(combined)
    combined = combined.rename(columns=rename_map)
    suffixed_cols = [v for k, v in rename_map.items() if v != k]
    print(f"\n  AllDBCombined: {len(combined)} rows × {len(combined.columns)} cols")
    if suffixed_cols:
        print(f"  Columns tagged with contributing DBs:")
        for new_name in suffixed_cols:
            print(f"    • {new_name}")

    # ── Write workbook ────────────────────────────────────────────────────────
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name="AllDBCombined", index=False)
        # Per-DB sheets: untouched copies of the source data sheet
        for n in (1, 2, 3, 4):
            raw_dfs[n].to_excel(writer, sheet_name=f"DB{n}", index=False)

        wb = writer.book
        header_font  = Font(bold=True, color="FFFFFF")
        sheet_fills  = {
            "AllDBCombined": PatternFill("solid", fgColor="263238"),  # near-black
            "DB1":           PatternFill("solid", fgColor="1565C0"),  # blue
            "DB2":           PatternFill("solid", fgColor="2E7D32"),  # green
            "DB3":           PatternFill("solid", fgColor="6A1B9A"),  # purple
            "DB4":           PatternFill("solid", fgColor="C62828"),  # red
        }
        for sheet_name, fill in sheet_fills.items():
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="center")

            # Apply date-only format to the Date column on AllDBCombined only.
            # The per-DB sheets stay untouched per user request.
            if sheet_name == "AllDBCombined":
                for col_cells in ws.iter_cols(min_row=1, max_row=1):
                    if col_cells[0].value == "Date":
                        letter = get_column_letter(col_cells[0].column)
                        for cell in ws[letter][1:]:
                            cell.number_format = "YYYY-MM-DD"

            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 50)

    # ── Console summary ───────────────────────────────────────────────────────
    print(f"\n{'─'*50}")
    print("[DONE] Combined workbook written:")
    print(f"  Path: {OUT_PATH}")
    print(f"  Sheets:")
    print(f"    AllDBCombined : {len(combined)} rows × {len(combined.columns)} cols (harmonized)")
    for n in (1, 2, 3, 4):
        print(f"    DB{n}           : {len(raw_dfs[n])} rows × {len(raw_dfs[n].columns)} cols (untouched)")

    print(f"\n  Total PIPL observed (sum TotalObserved): {int(total_pipl):,}")
    print(f"  Total banded count (sum TotalBanded):    {int(total_banded):,}")

    if warnings:
        print(f"\n  ⚠ {len(warnings)} warning(s):")
        for w in warnings[:20]:
            print(f"    - {w}")
        if len(warnings) > 20:
            print(f"    (and {len(warnings)-20} more — full list returned to caller)")
    else:
        print(f"\n  ✓ No species or harmonization warnings.")

    return warnings


if __name__ == "__main__":
    main()
